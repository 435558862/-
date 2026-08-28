import os
import pandas as pd
from typing import Dict, List, Optional, Union
from openpyxl import load_workbook
from PySide6.QtCore import QAbstractTableModel, QItemSelection, QItemSelectionModel, QModelIndex, Qt
from PySide6.QtGui import QAction, QFont, QKeyEvent
from PySide6.QtWidgets import (
    QAbstractItemView, QApplication, QFileDialog, QStyledItemDelegate, QTableWidgetItem, QMenu,
    QDialog, QHBoxLayout, QVBoxLayout, QLabel, QLineEdit, QComboBox, QCheckBox,
    QToolButton, QPushButton, QHeaderView, QMainWindow, QMessageBox, QTableView, QTableWidget
)
from src.gui.i18n import cell_zh, column_zh


class FindDialog(QDialog):
    """ Find dialog tool for excel-style tables. """

    def __init__(self, table: 'ExcelTable'):
        super().__init__(table)

        self._table = table

        # Declare placeholders.
        self._query_matches = []
        self._match_index = -1
        self._old_query = None
        self._old_scope = None
        self._old_exact = None

        # Declare UI placeholders.
        self._line_edit = None
        self._combo_scope = None
        self._checkbox_exact_match = None
        self._btn_prv = None
        self._btn_next = None
        self._btn_find = None
        self._label_status = None

        self._initialize_window()
        self._add_widgets()

    def _initialize_window(self):
        self.setWindowTitle('Find')

        # Allow dialog to float, while parent is active.
        self.setWindowModality(Qt.WindowModality.NonModal)
        self.setWindowFlag(Qt.WindowType.Tool, True)

    def _add_widgets(self):
        root = QVBoxLayout(self)

        # Row 1: query + scope + exact match
        r1 = QHBoxLayout()
        r1.addWidget(QLabel('Find:'))
        self._line_edit = QLineEdit()
        self._line_edit.setPlaceholderText('Search a keyword...')
        r1.addWidget(self._line_edit)

        self._combo_scope = QComboBox()
        self._combo_scope.addItems(['Entire Table'] + self._table.columns)
        self._combo_scope.currentIndexChanged.connect(self._find)
        r1.addWidget(self._combo_scope)

        self._checkbox_exact_match = QCheckBox('Exact match')
        self._checkbox_exact_match.toggled.connect(self._find)
        r1.addWidget(self._checkbox_exact_match)
        root.addLayout(r1)

        # Row 2: prev/next + status + refresh
        r2 = QHBoxLayout()
        self._label_status = QLabel('0 matches')
        self._btn_prev = QToolButton()
        self._btn_prev.setArrowType(Qt.ArrowType.LeftArrow)
        self._btn_prev.clicked.connect(self._goto_prev)
        self._btn_next = QToolButton()
        self._btn_next.setArrowType(Qt.ArrowType.RightArrow)
        self._btn_next.clicked.connect(self._goto_next)
        self._btn_find = QPushButton('Find')
        self._btn_find.clicked.connect(self._find)
        r2.addWidget(self._btn_prev)
        r2.addWidget(self._btn_next)
        r2.addStretch(1)
        r2.addWidget(self._label_status)
        r2.addStretch(1)
        r2.addWidget(self._btn_find)
        root.addLayout(r2)

        # Prefill from current selection.
        if hasattr(self._table, 'currentItem'):
            item = self._table.currentItem()
            if item is not None:
                self._line_edit.setText(item.text())
        else:
            index = self._table.currentIndex()
            if index.isValid():
                self._line_edit.setText(str(index.data() or ''))

    def _find(self):
        """ Searches the specified query in the table or specific columns. """

        # Fetch query and settings.
        raw = self._line_edit.text()
        query = raw if raw.strip() else None
        exact_match = self._checkbox_exact_match.isChecked()
        selected_scope_index = self._combo_scope.currentIndex()

        if query == '':
            return

        # Avoid finding the same query (if accidentally pressed with same settings).
        if query == self._old_query and exact_match == self._old_exact and selected_scope_index == self._old_scope:
            self._goto_next()
            return
        else:
            self._old_query = query
            self._old_exact = exact_match
            self._old_scope = selected_scope_index

        # Notify user whether they are performing expensive look-up procedure.
        if selected_scope_index == 0 and not exact_match:
            proceed = QMessageBox.warning(
                self,
                'Slow Operation',
                'Searching in the entire table without an exact match is an expensive operation. Do you want to proceed?',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            ) == QMessageBox.StandardButton.Yes
            if not proceed:
                return

        # Searching QTableWidget/QAbstractItemModel objects from a worker thread
        # is unsafe.  The large read-only league table uses a vectorized pandas
        # search, while editable tables are small enough to search directly.
        self._query_matches = self._table.compute_matches(
            query=query,
            exact_match=exact_match,
            selected_scope_index=selected_scope_index,
        )

        # Update obtained hit index.
        self._match_index = 0 if self._query_matches else -1

        # Highlight table items and update status label.
        self._table.highlight_items(self._query_matches)
        if self._query_matches:
            self._table.focus_item(self._query_matches[self._match_index])
        self._update_status()

    def _update_status(self):
        n = len(self._query_matches)

        if n == 0:
            self._label_status.setText('0 matches')
            self._btn_prev.setEnabled(False)
            self._btn_next.setEnabled(False)
        else:
            self._label_status.setText(f'{self._match_index + 1} / {n} matches')
            self._btn_prev.setEnabled(True)
            self._btn_next.setEnabled(True)

    def _goto_next(self):
        if self._query_matches:
            self._match_index = (self._match_index + 1) % len(self._query_matches)
            self._table.focus_item(self._query_matches[self._match_index])
            self._update_status()

    def _goto_prev(self):
        if self._query_matches:
            self._match_index = (self._match_index - 1) % len(self._query_matches)
            self._table.focus_item(self._query_matches[self._match_index])
            self._update_status()


class DataFrameTableModel(QAbstractTableModel):
    """Lazy read-only Qt model backed by a pandas DataFrame."""

    def __init__(self, df: pd.DataFrame, parent=None):
        super().__init__(parent)
        self._df = df.reset_index(drop=True).copy(deep=False)
        self.columns = self._df.columns.tolist()

    @property
    def dataframe(self) -> pd.DataFrame:
        return self._df

    @property
    def missing_rows(self) -> List[int]:
        return self._df.index[self._df.isna().any(axis=1)].tolist()

    def rowCount(self, parent=QModelIndex()) -> int:
        return 0 if parent.isValid() else self._df.shape[0]

    def columnCount(self, parent=QModelIndex()) -> int:
        return 0 if parent.isValid() else self._df.shape[1]

    def data(self, index: QModelIndex, role=Qt.ItemDataRole.DisplayRole):
        if not index.isValid():
            return None
        if role == Qt.ItemDataRole.DisplayRole:
            value = self._df.iat[index.row(), index.column()]
            if pd.isna(value):
                return ''
            return cell_zh(self.columns[index.column()], str(value))
        if role == Qt.ItemDataRole.TextAlignmentRole:
            return Qt.AlignmentFlag.AlignCenter
        return None

    def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
        if role != Qt.ItemDataRole.DisplayRole:
            return None
        if orientation == Qt.Orientation.Horizontal:
            return column_zh(self.columns[section])
        return str(section + 1)

    def sort(self, column: int, order=Qt.SortOrder.AscendingOrder):
        if column < 0 or column >= len(self.columns):
            return
        self.layoutAboutToBeChanged.emit()
        selected_column = self.columns[column]
        if selected_column == '赛事编号':
            # Preserve the official ticket-day semantics when the user clicks
            # this header. Plain string sorting interleaves 周五001/周六001 and
            # places 010 before 002 in some exports/views.
            from src.services.daily_sporttery import _sort_by_match_number
            self._df = _sort_by_match_number(self._df)
            if order == Qt.SortOrder.DescendingOrder:
                self._df = self._df.iloc[::-1].reset_index(drop=True)
        else:
            self._df = self._df.sort_values(
                by=selected_column,
                ascending=order == Qt.SortOrder.AscendingOrder,
                kind='mergesort',
                na_position='last',
            ).reset_index(drop=True)
        self.layoutChanged.emit()


class DataFrameTable(QTableView):
    """Fast read-only league table that renders only visible cells."""

    def __init__(
            self,
            parent: Optional[QMainWindow],
            df: pd.DataFrame,
            supports_sorting: bool = True,
            supports_query_search: bool = True,
    ):
        super().__init__(parent)
        self._parent = parent
        self.rows = df.shape[0]
        self.columns = df.columns.tolist()
        self._supports_query_search = supports_query_search
        self._hide_missing_rows = False
        self._hidden_missing_rows = []
        self._find_dialog = None
        self._table_model = DataFrameTableModel(df, self)
        self.setModel(self._table_model)

        self.setAlternatingRowColors(True)
        self.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectItems)
        self.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.setSortingEnabled(supports_sorting)

        header = self.horizontalHeader()
        header.setSectionsClickable(True)
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        header.setStretchLastSection(True)
        header.setDefaultSectionSize(110)
        font = QFont()
        font.setBold(True)
        header.setFont(font)

        self.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.customContextMenuRequested.connect(self._menu)
        self._table_model.layoutChanged.connect(self._refresh_hidden_rows)

    def _menu(self, pos):
        menu = QMenu(self)
        action_copy = QAction('Copy', self, triggered=self.copy_selection)
        action_copy.setToolTip('Copy the selected elements.')
        menu.addAction(action_copy)

        if self._supports_query_search:
            action_find = QAction('Find', self, triggered=self.open_find_dialog)
            action_find.setToolTip('Search a specific value in the table/column.')
            menu.addAction(action_find)

        header = self.horizontalHeader()
        is_auto = header.sectionResizeMode(0) == QHeaderView.ResizeMode.ResizeToContents
        action_resize = QAction('Auto-Resizing', self)
        action_resize.setCheckable(True)
        action_resize.setChecked(is_auto)
        action_resize.toggled.connect(
            lambda checked: header.setSectionResizeMode(
                QHeaderView.ResizeMode.ResizeToContents
                if checked else QHeaderView.ResizeMode.Interactive
            )
        )
        menu.addSeparator()
        menu.addAction(action_resize)
        menu.exec(self.viewport().mapToGlobal(pos))

    def copy_selection(self):
        indexes = self.selectedIndexes()
        if not indexes:
            if isinstance(self._parent, QMainWindow):
                self._parent.statusBar().showMessage('No text was copied.', 3000)
            return

        selected = {(index.row(), index.column()): str(index.data() or '') for index in indexes}
        rows = sorted({index.row() for index in indexes})
        columns = sorted({index.column() for index in indexes})
        lines = [
            '\t'.join(selected.get((row, column), '') for column in columns)
            for row in rows
        ]
        QApplication.clipboard().setText('\n'.join(lines))
        if isinstance(self._parent, QMainWindow):
            self._parent.statusBar().showMessage('Copied to clipboard.', 3000)

    def open_find_dialog(self):
        if self._find_dialog is None or not self._find_dialog.isVisible():
            self._find_dialog = FindDialog(self)
            self._find_dialog.show()
        else:
            self._find_dialog.activateWindow()
            self._find_dialog.raise_()

    def compute_matches(self, query: Optional[str], exact_match: bool, selected_scope_index: int) -> List[QModelIndex]:
        if not query:
            return []
        column_ids = range(len(self.columns)) if selected_scope_index <= 0 else [selected_scope_index - 1]
        normalized_query = str(query).casefold()
        hits = []
        hidden_rows = set(self._table_model.missing_rows) if self._hide_missing_rows else set()
        for column_id in column_ids:
            column_name = self.columns[column_id]
            displayed = self._table_model.dataframe.iloc[:, column_id].map(
                lambda value: '' if pd.isna(value) else cell_zh(column_name, str(value))
            )
            normalized = displayed.str.casefold()
            mask = normalized.eq(normalized_query) if exact_match else normalized.str.contains(normalized_query, regex=False)
            for row_id in mask.index[mask]:
                if row_id not in hidden_rows:
                    hits.append(self._table_model.index(int(row_id), column_id))
        hits.sort(key=lambda index: (index.row(), index.column()))
        return hits

    def highlight_items(self, items: List[QModelIndex]):
        self.clearSelection()
        if not items:
            return
        selection = QItemSelection()
        for index in items:
            selection.select(index, index)
        self.selectionModel().select(selection, QItemSelectionModel.SelectionFlag.Select)

    def focus_item(self, item: Optional[QModelIndex]):
        if item is not None and item.isValid():
            self.setCurrentIndex(item)
            self.scrollTo(item, QAbstractItemView.ScrollHint.PositionAtCenter)

    def hide_missing(self, hide: bool):
        self._hide_missing_rows = hide
        self._refresh_hidden_rows()

    def _refresh_hidden_rows(self):
        for row in self._hidden_missing_rows:
            self.setRowHidden(row, False)
        self._hidden_missing_rows = (
            self._table_model.missing_rows if self._hide_missing_rows else []
        )
        for row in self._hidden_missing_rows:
            self.setRowHidden(row, True)


class _TicketSortItem(QTableWidgetItem):
    """Table item whose display text and chronological sort key differ."""

    def __init__(self, text: str, sort_rank: int):
        super().__init__(text)
        self._sort_rank = sort_rank

    def __lt__(self, other):
        if isinstance(other, _TicketSortItem):
            return self._sort_rank < other._sort_rank
        return super().__lt__(other)


class ExcelTable(QTableWidget):
    """ Excel-style tables. Supports operations such as sorting, find, copy & paste. """

    def __init__(
            self,
            parent: Optional[QMainWindow],
            df: pd.DataFrame,
            readonly: bool = True,
            supports_sorting: bool = True,
            supports_query_search: bool = True,
            supports_deletion: bool = False
    ):
        super().__init__(df.shape[0], df.shape[1])

        self._parent = parent
        self.rows = df.shape[0]
        self.columns = df.columns.tolist()

        # Declare placeholders.
        self._supports_sorting = supports_sorting
        self._supports_query_search = supports_query_search
        self._supports_deletion = supports_deletion
        self._hide_missing_rows = False
        self._missing_rows = df.index[df.isna().any(axis=1)].tolist()
        self._missing_rows_set = set(self._missing_rows)
        self._hidden_rows_list = []

        # Declare UI placeholders.
        self._find_dialog = None
        self._initialize_table(df=df, readonly=readonly)

    def _initialize_table(self, df: pd.DataFrame, readonly: bool):
        """ Initializes and customizes table widget. """

        ticket_ranks = {}
        if '赛事编号' in df.columns:
            from src.services.daily_sporttery import _sort_by_match_number
            ranked = df.reset_index(drop=True).assign(_原始行=range(len(df)))
            ranked = _sort_by_match_number(ranked)
            ticket_ranks = {
                int(original): rank
                for rank, original in enumerate(ranked['_原始行'].tolist())
            }

        def add_data_to_table():
            for r in range(df.shape[0]):
                for c in range(df.shape[1]):
                    val = df.iat[r, c]
                    displayed = '' if pd.isna(val) else cell_zh(self.columns[c], str(val))
                    if self.columns[c] == '赛事编号':
                        item = _TicketSortItem(displayed, ticket_ranks.get(r, r))
                    else:
                        item = QTableWidgetItem(displayed)

                    # Center to column.
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.setItem(r, c, item)

                # Keep large league tables responsive while they are being
                # populated.  Qt widgets must only be modified on the GUI
                # thread; the previous TaskRunnerDialog worker caused random
                # freezes and high CPU usage under WSLg.
                if r and r % 250 == 0:
                    QApplication.processEvents()

        # Headers and fonts
        self.setHorizontalHeaderLabels([column_zh(column) for column in self.columns])
        hf = QFont()
        hf.setBold(True)
        self.horizontalHeader().setFont(hf)

        # Fill data: Centered (and optionally read-only).
        self.setSortingEnabled(False)   # Disable sorting before data-loading to increase re-size speed.
        self.setUpdatesEnabled(False)   # Disable updating before data-loading to increase re-size speed.

        # Add data on the GUI thread.  QTableWidget and QTableWidgetItem are
        # not thread-safe and must never be touched by a QThread worker.
        add_data_to_table()

        # Stylize table.
        self.setAlternatingRowColors(True)                              # Setting alternate colors between rows.
        self.setSelectionBehavior(self.SelectionBehavior.SelectItems)   # Enabling select.
        self.setSelectionMode(self.SelectionMode.ExtendedSelection)     # Enable multiple selection.

        if readonly:
            self.setEditTriggers(self.EditTrigger.NoEditTriggers)       # Setting read-only mode.

        # Enabling Sorting + Adjusting Header Sizing.
        self.resizeColumnsToContents()                                  # one-time auto-fit

        hh = self.horizontalHeader()
        hh.setSectionsClickable(True)
        hh.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)             # Manual default resizing policy.
        # hh.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)      # Too expensive on runtime.
        hh.setStretchLastSection(True)
        self.setUpdatesEnabled(True)

        if self._supports_sorting:
            self.setSortingEnabled(True)
            hh.setSortIndicatorShown(True)

        # Adding Shortcuts + Context Menu.
        self.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.customContextMenuRequested.connect(self._menu)

    def _menu(self, pos):
        """ Creates a default context menu for the table. """

        m = QMenu(self)
        action_copy = QAction('Copy', self, triggered=self.copy_selection)
        action_copy.setToolTip('Copy the selected elements.')
        m.addAction(action_copy)

        if self._supports_query_search:
            action_find = QAction('Find', self, triggered=self.open_find_dialog)
            action_find.setToolTip('Search a specific value in the table/column.')
            m.addAction(action_find)

        hh = self.horizontalHeader()
        is_auto = hh.sectionResizeMode(0) == QHeaderView.ResizeMode.ResizeToContents
        action_resize = QAction('Auto-Resizing', self)
        action_resize.setCheckable(True)
        action_resize.setChecked(is_auto)
        action_resize.toggled.connect(
            lambda checked: hh.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents if checked else QHeaderView.ResizeMode.Interactive)
        )
        action_resize.setToolTip('Whether to enable auto-resizing of the header widths. It might slow down performance.')
        m.addSeparator()
        m.addAction(action_resize)

        m.exec(self.viewport().mapToGlobal(pos))

    def keyPressEvent(self, event: QKeyEvent):
        key = event.key()
        if key and self._supports_deletion and key in (Qt.Key_Delete, Qt.Key_Backspace):
            for item in self.selectedItems():
                item.setText('')
        else:
            super().keyPressEvent(event)

    def copy_selection(self):
        """ Copies the selected table values into the clipboard. It is compatible with notepad/excel. """

        ranges = self.selectedRanges()
        if not ranges:
            if isinstance(self._parent, QMainWindow):
                self._parent.statusBar().showMessage('No text was copied.', 3000)
            return

        blocks = []
        for r in ranges:
            lines = []
            for row in range(r.topRow(), r.bottomRow() + 1):
                vals = []
                for col in range(r.leftColumn(), r.rightColumn() + 1):
                    it = self.item(row, col)
                    vals.append('' if it is None else it.text())
                lines.append('\t'.join(vals))
            blocks.append('\n'.join(lines))
        QApplication.clipboard().setText('\n'.join(blocks))

        if isinstance(self._parent, QMainWindow):
            self._parent.statusBar().showMessage('Copied to clipboard.', 3000)

    def open_find_dialog(self):
        """ Initializes/Shows the find dialog tool. """

        if self._find_dialog is None or not self._find_dialog.isVisible():
            self._find_dialog = FindDialog(self)
            self._find_dialog.show()
        else:
            self._find_dialog.activateWindow()
            self._find_dialog.raise_()

    def compute_matches(self, query: Optional[str], exact_match: bool, selected_scope_index: int) -> List[QTableWidgetItem]:
        """ Computes all query matches and returns all table entries where the query is found.
        :param query: The specified query to search.
        :param exact_match: Whether to search exact matches only.
        :param selected_col_index: The selected column to perform the search.
                             If selected_col == 'Entire Table', then the search is performed across all columns.
        """

        self.clearSelection()
        if not query:
            return []

        flags = Qt.MatchFlag.MatchFixedString if exact_match else Qt.MatchFlag.MatchContains

        if selected_scope_index <= 0:
            # Whole table: findItems searches all columns
            hits = self.findItems(query, flags)
        else:
            # Column scope: filter to that column
            all_hits = self.findItems(query, flags)
            col = selected_scope_index - 1
            hits = [it for it in all_hits if it.column() == col]

        # Sort hits by (row, col) ascending
        hits.sort(key=lambda it: (it.row(), it.column()))

        # Filter hits in hidden rows.
        if self._hide_missing_rows:
            hidden_set = self._missing_rows_set.union(self._hidden_rows_list)
            hits = [it for it in hits if it.row() not in hidden_set]

        # Select all in one shot to reduce selectionChanged churn
        sm = self.selectionModel()
        sm.blockSignals(True)
        for it in hits:
            it.setSelected(True)
        sm.blockSignals(False)
        return hits

    def highlight_items(self, items: List[QTableWidgetItem]):
        """ Highlights all selected items. """

        self.setUpdatesEnabled(False)

        self.clearSelection()
        for it in items:
            it.setSelected(True)

        self.setUpdatesEnabled(True)
        self.viewport().update()

    def highlight_rows(self, row_ids: List[int]):
        self.setUpdatesEnabled(False)

        self.clearSelection()

        # Remember original selection mode and behavior
        orig_behavior = self.selectionBehavior()
        orig_mode = self.selectionMode()

        self.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.setSelectionMode(QAbstractItemView.SelectionMode.MultiSelection)
        for r in row_ids:
            self.selectRow(r)
        self.setSelectionBehavior(orig_behavior)
        self.setSelectionMode(orig_mode)

        self.setUpdatesEnabled(True)
        self.viewport().update()

    def clear_selection(self):
        self.clearSelection()

    def focus_item(self, item: Optional[QTableWidgetItem]):
        """ Focuses a selected item in the table. """

        if item:
            self.setCurrentCell(item.row(), item.column())
            self.scrollToItem(item, QTableWidget.ScrollHint.PositionAtCenter)

    def hide_missing(self, hide: bool):
        """ Hides all rows that contain missing values. """

        self.setUpdatesEnabled(False)

        self._hide_missing_rows = hide
        for r in self._missing_rows:
            self.setRowHidden(r, hide)

        self.setUpdatesEnabled(True)
        self.viewport().update()

    def hide_columns(self, columns: Union[str, List[str]], hide: bool):
        """ Hides missing columns. """

        self.setUpdatesEnabled(False)

        if isinstance(columns, str):
            columns = [columns]

        for col in columns:
            self.setColumnHidden(self.columns.index(col), hide)

        self.setUpdatesEnabled(True)
        self.viewport().update()

    def set_new_hidden_rows(self, row_ids: Union[int, List[int]]):
        """ Un-hides previously hidden rows and hides new ones. """

        self.setUpdatesEnabled(False)

        # Un-hiding previously hidden rows.
        for r in self._hidden_rows_list:
            self.setRowHidden(r, False)

        # Hiding new rows.
        if isinstance(row_ids, int):
            row_ids = [row_ids]

        self._hidden_rows_list = row_ids
        for r in row_ids:
            self.setRowHidden(r, True)

        self.setUpdatesEnabled(True)
        self.viewport().update()

    def modify_columns(self, columns: List[str], data: List[List[str]], rows: Optional[List[int]] = None):
        self.setUpdatesEnabled(False)

        num_data_rows = len(data)

        # Validate data.
        if num_data_rows > self.rows or (rows is not None and num_data_rows > len(rows)):
            raise ValueError(f'Expected maximum {self.rows} rows, got {len(data)}.')

        col_ids = list(map(self.columns.index, columns))

        for i in range(len(data)):
            for j, c in enumerate(col_ids):
                item = QTableWidgetItem(str(data[i][j]))

                # Select 0,1,2... row from data if no rows are provided, else the rows[r].
                row_id = i if rows is None else rows[i]
                self.setItem(row_id, c, item)

        self.setUpdatesEnabled(True)
        self.viewport().update()


class StylizedTable:
    class ComboDelegate(QStyledItemDelegate):
        """ Delegate that shows a QComboBox when a cell is edited. """

        def __init__(self, options_provider, parent=None):
            super().__init__(parent)

            self._options_provider = options_provider

        def createEditor(self, parent, option, index):
            cb = QComboBox(parent)
            cb.setEditable(False)

            # Commit data immediately when a choice is made, then close
            cb.activated.connect(self._commit_and_close)
            return cb

        def _commit_and_close(self):
            """ Commits changes and closes combobox. """

            editor = self.sender()
            self.commitData.emit(editor)
            self.closeEditor.emit(editor, QStyledItemDelegate.EndEditHint.NoHint)

        def setEditorData(self, editor, index):
            """ Fills combo with options (can depend on index). """
            options = self._options_provider(index) or []
            editor.clear()
            editor.addItems(options)

            # Set current value to what's already in the cell.
            current = index.data(Qt.ItemDataRole.EditRole) or index.data(Qt.ItemDataRole.DisplayRole)
            if current is not None:
                i = editor.findText(str(current))
                if i >= 0:
                    editor.setCurrentIndex(i)

        def setModelData(self, editor, model, index):
            """ Writes selected text back to the model. """
            text = editor.currentText()
            model.setData(index, text, Qt.ItemDataRole.EditRole)
            model.setData(index, text, Qt.ItemDataRole.DisplayRole)

        def updateEditorGeometry(self, editor, option, index):
            editor.setGeometry(option.rect)

    def stylize_table(self, table: QTableWidget, options_dict: Dict[int, List[str]]) -> QTableWidget:
        """ Stylizes table with combobox options.
            :param table: The QTableWidget to be stylized.
            :param options_dict: A dict of {col_id: items} pairs, where col_id is the column index and items is a list
                                 of options. Unspecified columns will be modified manually via a line-edit.
            :returns: The stylized table.
        """

        def options_provider(index):
            options = options_dict.get(index.column(), [])
            return options

        delegate = self.ComboDelegate(options_provider=options_provider, parent=table)
        table.setEditTriggers(QAbstractItemView.EditTrigger.DoubleClicked | QAbstractItemView.EditTrigger.SelectedClicked)
        for col_id in options_dict:
            table.setItemDelegateForColumn(col_id, delegate)
        return table


class SimpleTableDialog(QDialog):
    def __init__(self, df: pd.DataFrame, parent: Optional[QMainWindow], title: str, readonly: bool = True):

        super().__init__(parent)

        self._df = df
        self.setWindowTitle(title)
        self.setWindowFlags(
            Qt.WindowType.Window |
            Qt.WindowType.WindowMinimizeButtonHint |
            Qt.WindowType.WindowCloseButtonHint
        )
        self.resize(600, 500)
        self.setWindowModality(Qt.WindowModality.NonModal)

        layout = QVBoxLayout(self)
        top = QHBoxLayout()
        top.addStretch()
        self.export_btn = QPushButton('Export')
        self.export_btn.setToolTip('Export DataFrame to (.csv). If file exists, append rows.')
        self.export_btn.clicked.connect(self._to_csv)
        top.addWidget(self.export_btn)
        layout.addLayout(top)

        self.table = ExcelTable(df=df, parent=self, readonly=readonly, supports_sorting=False, supports_query_search=False)
        layout.addWidget(self.table)

    def _to_csv(self):
        default_filepath = f'{self.windowTitle()}.csv'
        path, _ = QFileDialog.getSaveFileName(
            self,
            'Export to CSV',
            default_filepath,
            'CSV Files (*.csv);;Excel Files (*.xlsx)',
            options=QFileDialog.Option.DontConfirmOverwrite
        )

        if not path:
            return
        else:
            split_path = path.lower().rsplit('.', maxsplit=1)

            if len(split_path) == 1:
                path += '.csv'
                excel = False
            else:
                filetype = split_path[1]

                if filetype == 'csv':
                    excel = False
                elif filetype == 'xlsx':
                    excel = True
                else:
                    QMessageBox.critical(self, 'Export Failed', f'Invalid file type: "{filetype}"')
                    return

        try:
            if excel:
                if os.path.exists(path):
                    workbook = load_workbook(filename=path)
                    startrow, mode, header, if_sheet_exists = (workbook['Sheet1'].max_row, 'a', False, 'overlay')
                else:
                    startrow, mode, header, if_sheet_exists = (0, 'w', True, None)

                with pd.ExcelWriter(path=path, engine='openpyxl', mode=mode, if_sheet_exists=if_sheet_exists) as writer:
                    self._df.to_excel(
                        writer,
                        sheet_name='Sheet1',
                        startrow=startrow,
                        header=header,
                        index=False
                    )
            else:
                mode, header = ('a', False) if os.path.exists(path=path) else ('w', True)
                self._df.to_csv(path, mode=mode, header=header, index=False)

            QMessageBox.information(self, 'Success', 'Export Completed!')
        except Exception as e:
            QMessageBox.critical(self, 'Export Failed', f'Could not export data.\n\nError:\n{e}')
