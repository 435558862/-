import json
import os
from queue import Empty, Queue
import re
import sys
from threading import Thread
from datetime import date, datetime, timezone
from pathlib import Path

import pandas as pd
import numpy as np
from PySide6.QtCore import QTimer, Qt
from PySide6.QtGui import QBrush, QColor, QFont, QPainter, QPainterPath, QPalette, QPen
from PySide6.QtWidgets import (
    QComboBox, QDialog, QHBoxLayout, QLabel, QMessageBox, QPushButton,
    QStyle, QStyledItemDelegate, QVBoxLayout, QWidget,
)

from src.database.model import ModelDatabase
from src.gui.utils.taskrunner import TaskRunnerDialog
from src.gui.widgets.tables import ExcelTable
from src.network.fixtures.sporttery import SportteryMobileClient
from src.services.daily_learning import load_over_under_profile, review_and_learn
from src.services.lineups import lineup_api_configured, lineup_poll_interval_seconds
from src.services.daily_sporttery import (
    LEAGUE_ALIASES, _sort_by_match_number, backfill_missing_simulations,
    run_daily_sporttery,
)
from src.services.market_trends import build_trend_rows, live_snapshot_from_match, summarize_trend
from src.services.odds_tracking import (
    format_market_flow, read_odds_series, record_odds_snapshots,
    record_official_history,
)
from src.services.yesterday_review import _ticket_card_date, load_yesterday_hit_report
from src.services.value_selection import evaluate_value, historical_calibration


REPORT_ROOT = Path('storage/jingcai/reports')
PREDICTION_PATH = REPORT_ROOT / '最新竞彩预测.csv'
LEARNING_STATUS_PATH = Path('storage/jingcai/learning/status.json')
DAILY_RECOMMENDATION_ROOT = Path('storage/jingcai/daily_recommendations')
HALF_TIME_COMBINATION_ROOT = Path('storage/jingcai/half_time_combinations')
SETTLED_PREDICTIONS_PATH = Path('storage/jingcai/learning/settled_predictions.csv')
HALF_TIME_COMBINATION_STAKE = 1000.0
HALF_TIME_COMBINATION_GROUPS = {
    '胜': ('胜胜', '胜平', '胜负'),
    '平': ('平胜', '平平', '平负'),
    '负': ('负胜', '负平', '负负'),
}


def prediction_export_root() -> Path:
    """Return a writable, platform-native prediction export directory."""
    configured = os.environ.get('PROPHITBET_EXPORT_DIR', '').strip()
    if configured:
        return Path(configured).expanduser()
    if sys.platform in {'darwin', 'win32'}:
        return Path.home() / 'Desktop'
    wsl_desktop = Path('/mnt/c/Users/Administrator/Desktop')
    if wsl_desktop.is_dir():
        return wsl_desktop
    return Path.home()


def display_export_path(path: Path) -> str:
    """Format a native path for the completion message."""
    text = str(path)
    if sys.platform.startswith('linux') and text.startswith('/mnt/c/'):
        return text.replace('/mnt/c', 'C:', 1).replace('/', '\\')
    return text
ALL_MODELS = '__all__'
DEDICATED_MODELS = '__dedicated__'
GENERIC_MODELS = '__generic__'
SIMULATION_MODELS = '__simulation__'
INDEPENDENT_SIMULATION_SOURCE = '历史攻防双泊松蒙特卡洛'
DEDICATED_LEAGUE_COLUMN = '专用模型联赛'
REQUIRED_DEDICATED_MODELS = ('胜平负模型', '大小球模型', '比分模型', '半全场模型')


class _ComboPopupDelegate(QStyledItemDelegate):
    """Paint combo rows explicitly so desktop themes cannot force black selection."""

    def paint(self, painter, option, index):
        selected = bool(option.state & QStyle.StateFlag.State_Selected)
        hovered = bool(option.state & QStyle.StateFlag.State_MouseOver)
        background = '#2f79bd' if selected else '#dcebf8' if hovered else '#ffffff'
        foreground = '#ffffff' if selected else '#202020'
        painter.save()
        painter.fillRect(option.rect, QColor(background))
        painter.setPen(QColor(foreground))
        painter.drawText(
            option.rect.adjusted(6, 0, -4, 0),
            Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter,
            str(index.data(Qt.ItemDataRole.DisplayRole) or ''),
        )
        painter.restore()

    def sizeHint(self, option, index):
        size = super().sizeHint(option, index)
        size.setHeight(25)
        return size


class _WindowDragBar(QWidget):
    """In-window drag and close controls for platforms that hide dialog chrome."""

    def __init__(self, title: str, parent=None):
        super().__init__(parent)
        self._drag_offset = None
        self.setObjectName('windowDragBar')
        self.setFixedHeight(30)
        layout = QHBoxLayout(self)
        layout.setContentsMargins(8, 1, 3, 1)
        title_label = QLabel(title, self)
        title_label.setObjectName('windowDragTitle')
        layout.addWidget(title_label)
        layout.addStretch(1)
        close_button = QPushButton('✕', self)
        close_button.setObjectName('windowCloseButton')
        close_button.setToolTip('关闭')
        close_button.setFixedSize(28, 24)
        close_button.clicked.connect(lambda: self.window().close())
        layout.addWidget(close_button)

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            self._drag_offset = (
                event.globalPosition().toPoint()
                - self.window().frameGeometry().topLeft()
            )
            event.accept()
            return
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if (
            self._drag_offset is not None
            and event.buttons() & Qt.MouseButton.LeftButton
        ):
            self.window().move(event.globalPosition().toPoint() - self._drag_offset)
            event.accept()
            return
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        self._drag_offset = None
        super().mouseReleaseEvent(event)


class YesterdayHitDetailsDialog(QDialog):
    """A compact, cache-only view of yesterday's honest settlement details."""

    def __init__(self, parent=None):
        super().__init__(parent)
        frame, summary = load_yesterday_hit_report()
        title = (
            f'{summary["date"]} 最近已结算命中（昨日待补）'
            if summary.get('is_fallback') else f'{summary["date"]} 昨日命中明细'
        )
        self.setWindowModality(Qt.WindowModality.NonModal)
        self.setWindowFlags(
            Qt.WindowType.Window
            | Qt.WindowType.WindowMinimizeButtonHint
            | Qt.WindowType.WindowMaximizeButtonHint
            | Qt.WindowType.WindowCloseButtonHint
        )
        self.resize(1280, 620)

        root = QVBoxLayout(self)
        root.setContentsMargins(7, 7, 7, 7)
        root.setSpacing(5)
        headline = QLabel(summary['headline'])
        headline.setObjectName('yesterdayHeadline')
        headline.setWordWrap(True)
        root.addWidget(headline)
        if summary['patterns']:
            patterns = QLabel(f'规律：{summary["patterns"]}')
            patterns.setObjectName('yesterdayPatterns')
            patterns.setWordWrap(True)
            root.addWidget(patterns)
        table = ExcelTable(
            parent=self,
            df=frame,
            readonly=True,
            supports_sorting=True,
            supports_query_search=False,
        )
        self._highlight_hit_rows(table, frame)
        root.addWidget(table)
        self.setStyleSheet('''
            QLabel#yesterdayHeadline {
                color: #202020; font-size: 12px; font-weight: 600;
                padding: 4px 5px; background: #eef5fb;
                border: 1px solid #c8d9e8; border-radius: 3px;
            }
            QLabel#yesterdayPatterns {
                color: #404040; font-size: 12px; padding: 2px 5px;
            }
            QTableWidget {
                color: #202020; background: #ffffff;
                alternate-background-color: #f6f8fa;
                gridline-color: #cfd4d8; font-size: 12px;
            }
            QHeaderView::section {
                color: #202020; background: #e8edf1;
                border: 0; border-right: 1px solid #bdc5cb;
                border-bottom: 1px solid #9da5ab;
                padding: 2px 4px; font-size: 12px; font-weight: 600;
            }
        ''')

    @staticmethod
    def _highlight_hit_rows(table: ExcelTable, frame: pd.DataFrame) -> None:
        """Highlight only the market cells that actually hit."""
        foreground = QBrush(QColor('#c62828'))
        background = QBrush(QColor('#fff1f1'))
        hit_marks = ('（命中）', '（首中）', '（次中）', '（次1中）', '（次2中）',
                     '（冷中）', '（进中）')
        for row_index in range(table.rowCount()):
            for column_index in range(table.columnCount()):
                item = table.item(row_index, column_index)
                if item is not None and any(mark in item.text() for mark in hit_marks):
                    item.setForeground(foreground)
                    item.setBackground(background)
                    item.setToolTip('该项命中')


class _MarketTrendChart(QWidget):
    """Small dependency-free H/D/A implied-probability chart."""

    COLORS = {'H': '#ef5350', 'D': '#2f80d0', 'A': '#18a66a'}
    LABELS = {'H': '胜', 'D': '平', 'A': '负'}

    def __init__(self, parent=None):
        super().__init__(parent)
        self._rows = []
        self.setMinimumHeight(300)

    def set_rows(self, rows):
        self._rows = list(rows or [])
        self.update()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        painter.fillRect(self.rect(), QColor('#ffffff'))
        left, top, right, bottom = 48, 28, 18, 34
        width = max(1, self.width() - left - right)
        height = max(1, self.height() - top - bottom)
        painter.setFont(QFont('', 9))
        values = [
            float(row[key])
            for row in self._rows for key in ('H', 'D', 'A')
            if row.get(key) is not None
        ]
        minimum, maximum = 0.0, 1.0
        if values:
            span = max(values) - min(values)
            padding = max(0.012, span * 0.12)
            minimum = max(0.0, min(values) - padding)
            maximum = min(1.0, max(values) + padding)
            if maximum - minimum < 0.04:
                middle = (maximum + minimum) / 2
                minimum, maximum = max(0.0, middle - 0.02), min(1.0, middle + 0.02)
        value_span = max(0.0001, maximum - minimum)
        for step in range(6):
            value = minimum + value_span * step / 5
            y = top + height * (1.0 - (value - minimum) / value_span)
            painter.setPen(QPen(QColor('#dce3e9'), 1))
            painter.drawLine(left, int(y), left + width, int(y))
            painter.setPen(QColor('#66717a'))
            painter.drawText(4, int(y) - 8, 40, 16, Qt.AlignmentFlag.AlignRight,
                             f'{value:.0%}')
        legend_x = left
        for key in ('H', 'D', 'A'):
            painter.setPen(QPen(QColor(self.COLORS[key]), 3))
            painter.drawLine(legend_x, 12, legend_x + 18, 12)
            painter.setPen(QColor('#303840'))
            painter.drawText(legend_x + 23, 4, 30, 17,
                             Qt.AlignmentFlag.AlignLeft, self.LABELS[key])
            legend_x += 62
        if not self._rows:
            painter.setPen(QColor('#6b747c'))
            painter.drawText(self.rect(), Qt.AlignmentFlag.AlignCenter,
                             '正在等待实时赔率…')
            return
        denominator = max(1, len(self._rows) - 1)
        for key in ('H', 'D', 'A'):
            path = QPainterPath()
            for index, row in enumerate(self._rows):
                x = left + width * index / denominator
                y = top + height * (1.0 - (float(row[key]) - minimum) / value_span)
                if index == 0:
                    path.moveTo(x, y)
                else:
                    path.lineTo(x, y)
            painter.setPen(QPen(QColor(self.COLORS[key]), 2))
            painter.drawPath(path)
            painter.setBrush(QColor(self.COLORS[key]))
            for point_index, point_row in enumerate(self._rows):
                point_x = left + width * point_index / denominator
                point_y = top + height * (
                    1.0 - (float(point_row[key]) - minimum) / value_span
                )
                painter.drawEllipse(int(point_x) - 3, int(point_y) - 3, 6, 6)
            painter.setBrush(Qt.BrushStyle.NoBrush)
        label_indexes = sorted({0, len(self._rows) // 2, len(self._rows) - 1})
        painter.setPen(QColor('#66717a'))
        for index in label_indexes:
            raw_label = str(self._rows[index].get('label') or '')
            label = (
                f'初盘 {raw_label[-5:]}'
                if self._rows[index].get('is_opening') else raw_label[-8:]
            )
            x = left + width * index / denominator
            painter.drawText(int(x) - 38, top + height + 8, 76, 18,
                             Qt.AlignmentFlag.AlignCenter, label)


class MarketTrendDialog(QDialog):
    """Live H/D/A trend chart polling the official mobile odds feed."""

    def __init__(self, predictions: pd.DataFrame, parent=None):
        super().__init__(parent)
        self.setWindowTitle('实时市场走势图')
        self.setWindowModality(Qt.WindowModality.NonModal)
        self.resize(1080, 640)
        self._series = {
            key: list(rows) for key, rows in read_odds_series(
                max_rows_per_match=300, keep_opening=True,
            ).items()
        }
        self._live_queue = Queue()
        self._live_running = False
        self._did_select_live = False
        self._table = None

        root = QVBoxLayout(self)
        root.setContentsMargins(7, 7, 7, 7)
        root.setSpacing(5)
        controls = QHBoxLayout()
        controls.addWidget(QLabel('比赛'))
        self._match_selector = QComboBox()
        self._match_selector.setMinimumWidth(430)
        controls.addWidget(self._match_selector)
        controls.addWidget(QLabel('区间'))
        self._range_selector = QComboBox()
        for label, value in (('6小时', 6), ('12小时', 12), ('24小时', 24),
                             ('72小时', 72), ('全部', None)):
            self._range_selector.addItem(label, value)
        self._range_selector.setCurrentIndex(2)
        controls.addWidget(self._range_selector)
        controls.addStretch(1)
        root.addLayout(controls)

        self._live_status = QLabel('实时数据源：竞彩网官方移动端｜正在连接…')
        self._live_status.setObjectName('liveMarketStatus')
        self._live_status.setWordWrap(True)
        root.addWidget(self._live_status)
        self._summary = QLabel()
        self._summary.setObjectName('marketTrendSummary')
        self._summary.setWordWrap(True)
        root.addWidget(self._summary)
        self._chart = _MarketTrendChart(self)
        root.addWidget(self._chart, 3)
        self._table_container = QWidget(self)
        self._table_layout = QVBoxLayout(self._table_container)
        self._table_layout.setContentsMargins(0, 0, 0, 0)
        root.addWidget(self._table_container, 2)

        seen = set()
        for _, row in predictions.iterrows():
            value = row.get('比赛ID')
            try:
                match_id = str(int(value))
            except (TypeError, ValueError):
                match_id = str(value or '').strip()
            if not match_id or match_id in seen:
                continue
            seen.add(match_id)
            label = (
                f'{row.get("赛事编号", "")}  {row.get("联赛", "")}  '
                f'{row.get("主队", "")} vs {row.get("客队", "")}'
            )
            self._match_selector.addItem(label.strip(), match_id)
        if self._match_selector.count() == 0:
            for match_id, rows in self._series.items():
                latest = rows[-1] if rows else {}
                label = (
                    f'{latest.get("match_num", "")}  {latest.get("league", "")}  '
                    f'{latest.get("home", "")} vs {latest.get("away", "")}'
                )
                self._match_selector.addItem(label.strip(), match_id)

        self._match_selector.currentIndexChanged.connect(self._render)
        self._range_selector.currentIndexChanged.connect(self._render)
        self.setStyleSheet('''
            QLabel#marketTrendSummary {
                color: #1f2b34; background: #edf5fb;
                border: 1px solid #c2d8e8; border-radius: 3px;
                padding: 6px; font-size: 12px; font-weight: 600;
            }
            QComboBox {
                color: #202020; background: #ffffff;
                border: 1px solid #aeb5bb; border-radius: 3px;
                padding: 2px 20px 2px 5px;
            }
        ''')
        self._render()

        self._live_poll_timer = QTimer(self)
        self._live_poll_timer.setInterval(60_000)
        self._live_poll_timer.timeout.connect(self._start_live_poll)
        self._live_poll_timer.start()
        self._live_result_timer = QTimer(self)
        self._live_result_timer.setInterval(250)
        self._live_result_timer.timeout.connect(self._collect_live_poll)
        self._live_result_timer.start()
        QTimer.singleShot(0, self._start_live_poll)

    def _start_live_poll(self):
        if self._live_running:
            return
        self._live_running = True
        self._live_status.setText(
            '实时数据源：竞彩网官方移动端｜正在刷新…｜每60秒自动采集'
        )
        selected_id = str(self._match_selector.currentData() or '')
        Thread(
            target=self._fetch_live_market, args=(selected_id,), daemon=True,
        ).start()

    def _fetch_live_market(self, selected_id=''):
        captured_at = datetime.now(timezone.utc).isoformat(timespec='seconds')
        try:
            client = SportteryMobileClient(
                timeout=8.0, retries=1,
            )
            matches = client.selling_matches()
            history = (
                client.fixed_bonus_history(selected_id) if selected_id else {}
            )
            self._live_queue.put(('ok', captured_at, (matches, history, selected_id)))
        except Exception as error:
            self._live_queue.put(('error', captured_at, str(error)))

    def _collect_live_poll(self):
        try:
            state, captured_at, payload = self._live_queue.get_nowait()
        except Empty:
            return
        self._live_running = False
        shown_time = datetime.fromisoformat(
            captured_at.replace('Z', '+00:00'),
        ).astimezone().strftime('%H:%M:%S')
        if state == 'error':
            self._live_status.setText(
                f'实时数据暂不可用｜{shown_time}｜{str(payload)[:120]}'
            )
            return
        matches, official_history, history_match_id = payload
        persisted = record_odds_snapshots(
            matches, captured_at=captured_at,
        )
        history_added = record_official_history(
            history_match_id, official_history,
        ) if official_history and history_match_id else 0
        # Reload after both writes so the chart starts with the true official
        # opening row and includes all historical handicap prices immediately.
        self._series = {
            key: list(rows) for key, rows in read_odds_series(
                max_rows_per_match=300, keep_opening=True,
            ).items()
        }
        live_ids = set()
        first_live_id = None
        selector_ids = {
            str(self._match_selector.itemData(index) or '')
            for index in range(self._match_selector.count())
        }
        appended = 0
        unchanged = 0
        for raw in matches:
            snapshot = live_snapshot_from_match(raw, captured_at)
            if snapshot is None:
                continue
            match_id = snapshot['match_id']
            live_ids.add(match_id)
            if first_live_id is None:
                first_live_id = match_id
            if match_id not in selector_ids:
                match_number = (
                    raw.get('matchNumStr') or raw.get('matchNum') or ''
                )
                league = raw.get('leagueAllName') or raw.get('leagueAbbName') or ''
                home = raw.get('homeTeamAllName') or raw.get('homeTeamAbbName') or ''
                away = raw.get('awayTeamAllName') or raw.get('awayTeamAbbName') or ''
                label = f'{match_number}  {league}  {home} vs {away}'.strip()
                self._match_selector.addItem(label, match_id)
                selector_ids.add(match_id)
            rows = self._series.setdefault(match_id, [])
            if rows and self._same_market_snapshot(rows[-1], snapshot):
                unchanged += 1
            else:
                rows.append(snapshot)
                self._series[match_id] = rows[-300:]
                appended += 1
        selected_id = str(self._match_selector.currentData() or '')
        if not self._did_select_live and first_live_id:
            if selected_id not in live_ids:
                for index in range(self._match_selector.count()):
                    if str(self._match_selector.itemData(index) or '') == first_live_id:
                        self._match_selector.setCurrentIndex(index)
                        selected_id = first_live_id
                        break
            self._did_select_live = True
        selected_note = (
            '当前比赛已更新' if selected_id in live_ids
            else '当前比赛已停售或官方暂未返回'
        )
        self._live_status.setText(
            f'实时数据源：竞彩网官方移动端｜{shown_time}｜'
            f'补录官方初盘/变盘{history_added}点、入库{persisted}场、'
            f'新增走势点{appended}场、未变{unchanged}场｜'
            f'{selected_note}｜每60秒自动采集'
        )
        self._render()

    @staticmethod
    def _same_market_snapshot(left: dict, right: dict) -> bool:
        """Do not draw artificial horizontal segments for unchanged odds."""
        keys = ('had', 'hhad', 'ttg')
        return all((left.get(key) or {}) == (right.get(key) or {}) for key in keys)

    def closeEvent(self, event):
        self._live_poll_timer.stop()
        self._live_result_timer.stop()
        super().closeEvent(event)

    def _render(self):
        match_id = self._match_selector.currentData()
        hours = self._range_selector.currentData()
        rows = build_trend_rows(match_id, self._series, hours=hours)
        summary = summarize_trend(rows)
        probabilities = summary.get('latest_probabilities') or {}
        probability_text = ' / '.join(
            f'{label}{probabilities.get(key, 0):.1%}'
            for key, label in (('H', '胜'), ('D', '平'), ('A', '负'))
        ) if probabilities else '--'
        self._summary.setText(
            f'综合结论：{summary["conclusion"]}　｜　'
            f'{summary["handicap"]}　｜　{summary["total_goals"]}　｜　'
            f'稳定性：{summary["stability"]}　｜　实时曲线点：{summary["observations"]}个\n'
            f'当前胜平负：{probability_text}　｜　'
            f'初盘基准：{rows[0]["label"] if rows else "--"}'
        )
        self._chart.set_rows(rows)
        frame = pd.DataFrame([{
            '记录时间': row['label'],
            '胜赔': f'{row["had_H"]:.2f}',
            '平赔': f'{row["had_D"]:.2f}',
            '负赔': f'{row["had_A"]:.2f}',
            '胜概率': f'{row["H"]:.1%}', '平概率': f'{row["D"]:.1%}',
            '负概率': f'{row["A"]:.1%}',
            '让球线': '' if row['hhad_line'] is None else f'{row["hhad_line"]:g}',
            '让球胜赔': '' if row['hhad_H'] is None else f'{row["hhad_H"]:.2f}',
            '让球平赔': '' if row['hhad_D'] is None else f'{row["hhad_D"]:.2f}',
            '让球负赔': '' if row['hhad_A'] is None else f'{row["hhad_A"]:.2f}',
            '大球': '' if row['over'] is None else f'{row["over"]:.1%}',
            '小球': '' if row['under'] is None else f'{row["under"]:.1%}',
        } for row in rows])
        if self._table is None:
            self._table = ExcelTable(
                parent=self, df=frame, readonly=True, supports_sorting=False,
                supports_query_search=False,
            )
            self._table_layout.addWidget(self._table)
        else:
            self._table.update_dataframe(frame)
        header = self._table.horizontalHeader()
        header.setStretchLastSection(False)
        widths = (190, 55, 55, 55, 65, 65, 65, 62, 76, 76, 76, 72, 72)
        for column_index, width in enumerate(widths):
            if column_index < self._table.columnCount():
                self._table.setColumnWidth(column_index, width)
        self._table.verticalHeader().setDefaultSectionSize(26)


def _safe_filename(value: str) -> str:
    return re.sub(r'[\\/:*?"<>|]+', '-', str(value)).strip(' .-') or '当前模型'


def write_predictions_xlsx(frame: pd.DataFrame, path: Path):
    """Write a Windows-ready workbook with readable headers and filters."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        frame.to_excel(writer, sheet_name='竞彩预测', index=False)
        sheet = writer.sheets['竞彩预测']
        sheet.freeze_panes = 'A2'
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(color='FFFFFF', bold=True)
            cell.fill = PatternFill('solid', fgColor='1F4E78')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for index, column in enumerate(frame.columns, start=1):
            values = [str(column), *(str(value) for value in frame[column].fillna('').head(500))]
            sheet.column_dimensions[get_column_letter(index)].width = min(
                36, max(10, max(map(len, values)) + 2),
            )
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical='center')
                if str(cell.value or '').startswith('★'):
                    cell.font = Font(color='C62828', bold=True)
                    cell.fill = PatternFill('solid', fgColor='FFF1F1')


def trained_dedicated_leagues() -> list[str]:
    """List leagues with the complete model set used by daily predictions."""
    if not LEAGUE_ALIASES:
        return []
    try:
        probe = ModelDatabase(next(iter(LEAGUE_ALIASES)))
        index = probe.index
    except Exception:
        # A damaged/temporarily unavailable index must not prevent the daily
        # prediction window from opening; report-derived options still work.
        return []
    result = []
    for league in LEAGUE_ALIASES:
        model_ids = set(index.get(league, {}))
        required = {f'{league}{suffix}' for suffix in REQUIRED_DEDICATED_MODELS}
        if required.issubset(model_ids):
            result.append(league)
    return result


def _prediction_model_scopes(df: pd.DataFrame) -> pd.Series:
    """Return the dedicated league used by every row, or an empty scope."""
    if DEDICATED_LEAGUE_COLUMN in df.columns:
        return df[DEDICATED_LEAGUE_COLUMN].fillna('').astype(str)
    # Backward compatibility for reports created before model scope was saved.
    basis = df.get('预测依据', pd.Series('', index=df.index)).fillna('').astype(str)
    leagues = df.get('联赛', pd.Series('', index=df.index)).fillna('').astype(str)
    return leagues.where(basis.eq('历史数据训练模型'), '')


def _upcoming_predictions(
        predictions: pd.DataFrame,
        now: datetime | None = None,
) -> pd.DataFrame:
    """Keep only fixtures that have not kicked off; retain malformed dates visibly."""
    if predictions.empty or '比赛时间' not in predictions.columns:
        return predictions.copy()
    kickoff = pd.to_datetime(predictions['比赛时间'], errors='coerce')
    return predictions.loc[
        kickoff.isna() | kickoff.gt(now or datetime.now())
    ].copy()


def _score_recommendation_mask(predictions: pd.DataFrame) -> pd.Series:
    """Rows whose audited top-score probability reaches the 12% gate."""
    status = predictions.get(
        '比分推荐状态', pd.Series('', index=predictions.index),
    ).fillna('').astype(str)
    probability = pd.to_numeric(
        predictions.get(
            '原始最高概率比分概率', pd.Series(float('nan'), index=predictions.index),
        ),
        errors='coerce',
    )
    return status.eq('推荐') | (status.eq('') & probability.ge(0.12))


def _daily_priority_aspects(predictions: pd.DataFrame) -> pd.Series:
    """Map the unified daily recommendation list back onto main-table cells."""
    aspects = pd.Series([[] for _ in range(len(predictions))], index=predictions.index)
    if predictions.empty:
        return aspects
    recommendations = build_daily_recommendations(predictions, future_only=False)
    if '推荐等级' in recommendations.columns:
        recommendations = recommendations.loc[
            recommendations['推荐等级'].eq('核心重点')
        ].copy()
    labels = {
        '胜平负': '胜负', '让球胜平负': '让球', '总进球': '总进球',
        '比分': '比分', '半全场': '半全场',
    }
    for index, row in predictions.iterrows():
        number = str(row.get('赛事编号') or '')
        matches = recommendations.loc[
            recommendations['赛事编号'].astype(str).eq(number), '推荐玩法',
        ]
        aspects.at[index] = [labels[value] for value in matches if value in labels]
    return aspects


def _legacy_daily_priority_aspects(predictions: pd.DataFrame) -> pd.Series:
    """Legacy threshold audit retained for historical analysis only."""
    aspects = pd.Series([[] for _ in range(len(predictions))], index=predictions.index)
    if predictions.empty:
        return aspects
    match_times = predictions.get(
        '比赛时间', pd.Series('', index=predictions.index),
    )
    match_numbers = predictions.get(
        '赛事编号', pd.Series('', index=predictions.index),
    )
    days = pd.Series([
        card_day.isoformat() if card_day is not None else '全部'
        for card_day in (
            _ticket_card_date(match_time, match_number)
            for match_time, match_number in zip(match_times, match_numbers)
        )
    ], index=predictions.index)
    gate = predictions.get(
        '盘口门控', pd.Series('', index=predictions.index),
    ).fillna('').astype(str)
    stable = ~gate.str.contains('冲突|震荡|不稳定', regex=True)

    def numbers(column: str) -> pd.Series:
        return pd.to_numeric(
            predictions.get(column, pd.Series(float('nan'), index=predictions.index)),
            errors='coerce',
        )

    advice = predictions.get(
        '建议状态', pd.Series('', index=predictions.index),
    ).fillna('').astype(str)
    ou_profile = load_over_under_profile() or {
        'directions': [
            {'pick': '大于2.5球', 'enabled': True, 'threshold': 0.60},
            {'pick': '小于2.5球', 'enabled': False, 'threshold': 0.75},
        ],
    }
    ou_rules = {
        str(row.get('pick')): row for row in ou_profile.get('directions', [])
    }
    ou_pick = predictions.get(
        '大小球首选', pd.Series('', index=predictions.index),
    ).fillna('').astype(str)
    ou_threshold = ou_pick.map(
        lambda pick: float((ou_rules.get(pick) or {}).get('threshold', 1.0)),
    )
    ou_enabled = ou_pick.map(
        lambda pick: bool((ou_rules.get(pick) or {}).get('enabled', False)),
    )
    candidates = {
        '胜负': (
            advice.isin(('精选主推', '高置信主推'))
            & numbers('胜平负首选概率').ge(0.625) & stable,
            numbers('胜平负首选概率'),
        ),
        # Draws need their own calibrated gate: a universal 62.5% result gate
        # structurally excludes them.  On sealed local settlements, >=32%
        # produced 4/7 hits; looser 28-30% gates fell below 38%.
        '平局': (
            numbers('模型平局概率').ge(0.32) & stable,
            numbers('模型平局概率'),
        ),
        # Settled directional audit: -1/让胜 is 10/14 (71.4%) and
        # +1/让负 is 4/6 (66.7%). The former probability/edge-only selector
        # picked the wrong structures and went 0/7 as a daily priority.
        '让球': (
            (
                (numbers('官方让球数').eq(-1)
                 & predictions.get('让球首选', pd.Series('', index=predictions.index)).eq('胜'))
                | (numbers('官方让球数').eq(1)
                   & predictions.get('让球首选', pd.Series('', index=predictions.index)).eq('负'))
            )
            & numbers('让球首选概率').ge(0.38) & stable,
            numbers('让球首选概率'),
        ),
        '大小球': (
            ou_enabled & numbers('大小球首选概率').ge(ou_threshold) & stable,
            numbers('大小球首选概率'),
        ),
        '半全场': (
            numbers('半全场首选概率').ge(0.35) & stable,
            numbers('半全场首选概率'),
        ),
        # Keep exact score selective but visible: choose only the strongest
        # audited score per day. Requiring 15% suppressed every current match,
        # while the score model's established display gate is 12%.
        '比分': (
            _score_recommendation_mask(predictions)
            & numbers('原始最高概率比分概率').ge(0.12) & stable,
            numbers('原始最高概率比分概率'),
        ),
    }
    for _, group_indices in days.groupby(days, sort=False).groups.items():
        group_indices = list(group_indices)
        for label, (eligible, strength) in candidates.items():
            available = [index for index in group_indices if bool(eligible.loc[index])]
            if not available:
                continue
            best = strength.loc[available].idxmax()
            aspects.at[best] = [*aspects.at[best], label]
    return aspects


def _mark_priority_cells(
        display: pd.DataFrame, priorities: pd.Series,
) -> dict[int, list[str]]:
    """Mark only the exact market cells selected for daily emphasis."""
    columns = {
        '胜负': '综合方向',
        '平局': '综合方向',
        '让球': '让球',
        '大小球': '大小球', '总进球': '总进球',
        '半全场': '半全场',
        '比分': '比分',
    }
    marked: dict[int, list[str]] = {}
    for row_index, labels in priorities.items():
        for label in labels:
            column = columns.get(label)
            if column not in display.columns:
                continue
            value = str(display.at[row_index, column] or '')
            display.at[row_index, column] = f'★{label}重点｜{value}'
            marked.setdefault(row_index, []).append(column)
    return marked


def _priority_summary(predictions: pd.DataFrame) -> str:
    """Return compact counts for recommendations that actually pass gates."""
    counts: dict[str, int] = {}
    for labels in _daily_priority_aspects(predictions):
        for label in labels:
            counts[label] = counts.get(label, 0) + 1
    ordered = ('胜负', '让球', '总进球', '半全场', '比分')
    selected = [f'{label}{counts[label]}' for label in ordered if counts.get(label)]
    return '今日重点 ' + ('·'.join(selected) if selected else '暂无达标项')


def build_daily_recommendations(
        predictions: pd.DataFrame, future_only: bool = True,
) -> pd.DataFrame:
    """Select at most eight risk-adjusted fixtures per card day.

    The formal model owns the pick.  Market movement and the independent
    Monte Carlo model are vetoes, never alternative sources of a pick.  Exact
    score and half/full remain visible in the main table but are deliberately
    excluded from the high-hit-rate daily list.
    """
    columns = [
        '比赛日期', '赛事编号', '联赛', '对阵', '推荐玩法', '重点选项',
        '最佳比分', '高倍候选',
        '推荐等级', '正式主模型', '正式模型概率', '价值评估', '建议仓位',
        '盘口验证', '蒙特卡洛是否同向', '阵容验证',
        '比分参考', '半全场参考', '入选理由',
    ]
    if predictions.empty:
        return pd.DataFrame(columns=columns)
    source = _upcoming_predictions(predictions) if future_only else predictions.copy()
    active = _sort_by_match_number(source).reset_index(drop=True)
    allow_observation_conflicts = len(active) >= 6

    def number(row: pd.Series, column: str) -> float:
        value = pd.to_numeric(row.get(column), errors='coerce')
        return float(value) if pd.notna(value) else float('nan')

    def first_simulation_pick(row: pd.Series, market: str) -> str:
        column = {
            '胜平负': '模拟胜负', '让球胜平负': '模拟让球',
        }[market]
        text = str(row.get(column) or '').strip()
        return re.split(r'[\s/｜]', text, maxsplit=1)[0]

    def implied(row: pd.Series, columns_: tuple[str, str, str]) -> np.ndarray | None:
        odds = np.array([number(row, column) for column in columns_], dtype=float)
        if not np.isfinite(odds).all() or np.any(odds <= 1.0):
            return None
        inverse = 1.0 / odds
        return inverse / inverse.sum()

    def formal_margin(row: pd.Series, columns_: tuple[str, str, str]) -> float:
        values = np.array([number(row, column) for column in columns_], dtype=float)
        values = values[np.isfinite(values)]
        if len(values) < 2:
            return 0.0
        values.sort()
        return float(values[-1] - values[-2])

    def probability_text(row: pd.Series, column: str) -> str:
        value = number(row, column)
        return f'{value:.1%}' if np.isfinite(value) else ''

    def score_reference(row: pd.Series) -> str:
        picks = []
        for pick_column, probability_column in (
                ('首选比分', '首选比分概率'),
                ('次选比分', '次选比分概率'),
                ('第三比分', '第三比分概率'),
        ):
            pick = str(row.get(pick_column) or '').strip()
            if not pick or pick.lower() == 'nan' or pick in picks:
                continue
            probability = probability_text(row, probability_column)
            picks.append(f'{pick}（{probability}）' if probability else pick)
        return ' / '.join(picks[:3]) or '暂无可靠比分'

    def best_score(row: pd.Series) -> str:
        pick = str(row.get('首选比分') or '').strip()
        if not pick or pick.lower() == 'nan':
            return '—'
        probability = number(row, '首选比分概率')
        return (
            f'◎ {pick}（{probability:.1%}）'
            if np.isfinite(probability) else f'◎ {pick}'
        )

    def high_odds_reference(row: pd.Series) -> str:
        """Return one auditable high-price candidate, never a core pick.

        A high SP alone is not enough. The formal probability must leave at
        least a small raw edge, while a conservative probability haircut may
        not make the candidate severely negative. This intentionally yields
        no candidate on many fixtures rather than manufacturing a long shot.
        """
        candidates = []

        def consider(
                market: str, label: str, probability_column: str,
                odds_column: str, haircut: float,
        ) -> None:
            probability = number(row, probability_column)
            odds = number(row, odds_column)
            if (
                not np.isfinite(probability) or not np.isfinite(odds)
                or probability < 0.18 or odds < 2.80
            ):
                return
            raw_ev = probability * odds - 1.0
            conservative_ev = max(0.01, probability - haircut) * odds - 1.0
            if raw_ev < 0.05 or conservative_ev < -0.08:
                return
            candidates.append((
                conservative_ev, raw_ev, probability, odds, market, label,
            ))

        for label, probability_column, odds_column in zip(
                ('胜', '平', '负'),
                ('模型主胜概率', '模型平局概率', '模型客胜概率'),
                ('官方胜奖金', '官方平奖金', '官方负奖金')):
            consider('胜平负', label, probability_column, odds_column, 0.025)
        line = number(row, '官方让球数')
        if np.isfinite(line):
            for label, probability_column, odds_column in zip(
                    ('让胜', '让平', '让负'),
                    ('模型让胜概率', '模型让平概率', '模型让负概率'),
                    ('官方让胜奖金', '官方让平奖金', '官方让负奖金')):
                consider(
                    '让球', f'{line:+g}球 {label}',
                    probability_column, odds_column, 0.040,
                )
        if not candidates:
            return '—'
        _, raw_ev, probability, odds, market, label = max(candidates)
        return (
            f'◆ {market}·{label}（SP {odds:.2f}｜模型{probability:.1%}｜'
            f'优势{raw_ev:+.1%}｜高风险）'
        )

    def half_full_reference(row: pd.Series) -> str:
        picks = []
        for pick_column, probability_column in (
                ('半全场首选', '半全场首选概率'),
                ('半全场次选', '半全场次选概率'),
        ):
            pick = str(row.get(pick_column) or '').strip()
            if not pick or pick.lower() == 'nan' or pick in picks:
                continue
            probability = probability_text(row, probability_column)
            picks.append(f'{pick}（{probability}）' if probability else pick)
        return ' / '.join(picks) or '暂无可靠半全场'

    def market_support(row: pd.Series, market: str, formal_pick: str) -> tuple[bool, str]:
        gate = str(row.get('盘口门控') or '')
        if any(word in gate for word in ('冲突', '震荡', '不稳定')):
            return False, gate or '盘口不稳定'
        codes = {'胜': 0, '平': 1, '负': 2}
        pick_index = codes.get(formal_pick)
        if pick_index is None:
            return False, '正式方向无法核验'
        if market == '胜平负':
            opening_columns = ('首次采集胜奖金', '首次采集平奖金', '首次采集负奖金')
            current_columns = ('官方胜奖金', '官方平奖金', '官方负奖金')
        else:
            opening_line = number(row, '首次采集让球数')
            current_line = number(row, '官方让球数')
            if not np.isfinite(opening_line) or not np.isfinite(current_line):
                return False, '缺少让球线快照'
            if not np.isclose(opening_line, current_line):
                return False, f'让球线变化 {opening_line:+g}→{current_line:+g}'
            opening_columns = ('首次采集让胜奖金', '首次采集让平奖金', '首次采集让负奖金')
            current_columns = ('官方让胜奖金', '官方让平奖金', '官方让负奖金')
        opening = implied(row, opening_columns)
        current = implied(row, current_columns)
        if opening is None or current is None:
            return False, '缺少初盘或当前盘口'
        if int(np.argmax(current)) != pick_index:
            return False, '当前盘口首选与正式模型不同向'
        movement = float(current[pick_index] - opening[pick_index])
        if movement < -0.015:
            return False, f'临场明显反向 {movement:+.1%}'
        if movement < -0.005:
            return False, f'正式方向走弱 {movement:+.1%}'
        if movement > 0.015:
            state = '持续走强'
        elif abs(movement) <= 0.003:
            state = '当前支持·去水概率基本不变'
        else:
            state = '当前支持·轻微走强'
        return True, f'{state} {movement:+.1%}'

    candidates_by_day: dict[str, list[dict]] = {}
    for _, row in active.iterrows():
        card_day = _ticket_card_date(row.get('比赛时间'), row.get('赛事编号'))
        day_text = card_day.isoformat() if card_day is not None else str(row.get('比赛时间') or '')[:10]
        market_candidates: list[tuple[str, str, str, float, float, float]] = []

        def add(
                market: str, choice: str, formal_pick: str, probability: float,
                probability_columns: tuple[str, str, str],
                odds_columns: tuple[str, str, str],
        ) -> None:
            if not choice or not np.isfinite(probability) or probability <= 0:
                return
            pick_index = {'胜': 0, '平': 1, '负': 2}.get(formal_pick)
            if pick_index is None:
                return
            odds = number(row, odds_columns[pick_index])
            if not np.isfinite(odds) or odds <= 1.0:
                return
            margin = formal_margin(row, probability_columns)
            market_candidates.append((
                market, choice, formal_pick, probability, margin, odds,
            ))

        regular_offered = all(np.isfinite(number(row, column)) for column in (
            '官方胜奖金', '官方平奖金', '官方负奖金',
        ))
        if regular_offered:
            add(
                '胜平负', str(row.get('胜平负首选') or ''),
                str(row.get('胜平负首选') or ''),
                number(row, '胜平负首选概率'),
                ('模型主胜概率', '模型平局概率', '模型客胜概率'),
                ('官方胜奖金', '官方平奖金', '官方负奖金'),
            )
        line = number(row, '官方让球数')
        handicap_offered = np.isfinite(line) and all(
            np.isfinite(number(row, column)) for column in (
                '官方让胜奖金', '官方让平奖金', '官方让负奖金',
            )
        )
        if handicap_offered:
            handicap_pick = str(row.get('让球首选') or '').strip()
            add(
                '让球胜平负', f'{line:+g}球 {handicap_pick}', handicap_pick,
                number(row, '让球首选概率'),
                ('模型让胜概率', '模型让平概率', '模型让负概率'),
                ('官方让胜奖金', '官方让平奖金', '官方让负奖金'),
            )
        for market, choice, formal_compare, probability, margin, official_odds in market_candidates:
            monte_pick = first_simulation_pick(row, market)
            monte_compare = monte_pick.removeprefix('让') if market == '让球胜平负' else monte_pick
            if not monte_pick:
                continue
            monte_aligned = formal_compare == monte_compare
            monte_conflict = not monte_aligned
            supported, support_text = market_support(row, market, formal_compare)
            if not supported:
                continue
            lineup_status = str(row.get('首发状态') or '')
            lineup_conflict = bool(row.get('阵容方向冲突'))
            lineup_warning = str(row.get('阵容预警级别') or '无')
            if lineup_conflict or lineup_warning == '高':
                continue
            empirical_accuracy = number(
                row,
                '让球历史命中率' if market == '让球胜平负' else '同阈值历史命中率',
            )
            empirical_samples = number(
                row,
                '让球回测样本数' if market == '让球胜平负' else '筛选回测样本数',
            )
            if not np.isfinite(empirical_accuracy) or not np.isfinite(empirical_samples):
                learned_accuracy, learned_samples = historical_calibration(
                    market, probability,
                )
                if learned_accuracy is not None:
                    empirical_accuracy = learned_accuracy
                    empirical_samples = float(learned_samples)
            decision = evaluate_value(
                market, probability, official_odds,
                empirical_accuracy=(
                    empirical_accuracy if np.isfinite(empirical_accuracy) else None
                ),
                empirical_samples=(
                    int(empirical_samples) if np.isfinite(empirical_samples) else 0
                ),
            )
            handicap_observation = (
                market == '让球胜平负'
                and probability >= 0.55 and not decision.promoted
            )
            # Keep the three-way cross-check intact, but do not let a strict
            # value gate reduce a usable daily card to one row. These rows
            # are explicitly observations (no stake) and are eligible only
            # when the formal direction, current market, and Monte Carlo are
            # already aligned. The value gate still controls all starred
            # picks and suggested stakes.
            fallback_observation = (
                not decision.promoted
                and not handicap_observation
                and probability >= 0.45
                and margin >= 0.08
            )
            # A Monte Carlo disagreement must never become a starred pick or
            # a suggested stake. When the daily card would otherwise be too
            # sparse, retain it as a clearly labelled observation so the user
            # can see the conflict and decide independently.
            monte_observation = (
                allow_observation_conflicts
                and monte_conflict
                and not decision.promoted
                and not handicap_observation
                and probability >= 0.45
                and margin >= 0.08
            )
            if not monte_aligned and not monte_observation:
                continue
            if not decision.promoted and not handicap_observation and not fallback_observation and not monte_observation:
                continue
            lineup_text = (
                f'已确认·预警{lineup_warning}' if lineup_status == '已确认'
                else '未确认·不调整模型'
            )
            display_grade = (
                '盘口观察' if handicap_observation else
                decision.grade if decision.promoted else
                '蒙特反向观察' if monte_observation else '综合观察'
            )
            grade_rank = 2 if display_grade == '核心重点' else (
                1 if display_grade == '可买优选' else 0
            )
            quality = (
                grade_rank * 10.0 + decision.conservative_ev * 5.0
                + probability + margin
            )
            candidates_by_day.setdefault(day_text, []).append({
                '_quality': quality,
                '比赛日期': day_text,
                '赛事编号': row.get('赛事编号', ''),
                '联赛': row.get('联赛', ''),
                '对阵': f'{row.get("主队", "")} vs {row.get("客队", "")}',
                '推荐玩法': market,
                '重点选项': (
                    f'· {choice}'
                    if handicap_observation or fallback_observation or monte_observation
                    else f'★ {choice}'
                ),
                '最佳比分': best_score(row),
                '高倍候选': high_odds_reference(row),
                '推荐等级': display_grade,
                '正式主模型': str(row.get('胜负模型类别') or '市场基线'),
                '正式模型概率': f'{probability:.1%}',
                '价值评估': (
                    f'SP {official_odds:.2f}｜保守概率 {decision.conservative_probability:.1%}'
                    f'｜EV {decision.conservative_ev:+.1%}'
                ),
                '建议仓位': (
                    f'≤{decision.stake_fraction:.1%}本金'
                    if decision.stake_fraction > 0 else '不投注'
                ),
                '盘口验证': support_text,
                '蒙特卡洛是否同向': (
                    f'同向（蒙特：{monte_pick}）'
                    if monte_aligned else f'反向观察（蒙特：{monte_pick}）'
                ),
                '阵容验证': lineup_text,
                '比分参考': score_reference(row),
                '半全场参考': half_full_reference(row),
                '入选理由': (
                    (
                        '正式模型与盘口同向，但蒙特反向，仅展示观察方向'
                        if monte_observation else
                        '三方同向但保守EV未达标，仅展示观察方向'
                        if handicap_observation or fallback_observation else
                        f'{display_grade}；正式模型定方向；领先第二方向{margin:.1%}；'
                        '保守EV达标；盘口与独立模拟双重同向'
                    )
                ),
            })
    # Rank one recommendation per fixture across the whole daily card. The
    # card target is 6–8 rows; if fewer than six survive the three-way gate,
    # the honest result is fewer rows rather than inventing a direction.
    ranked = sorted(
        (item for day_rows in candidates_by_day.values() for item in day_rows),
        key=lambda item: item['_quality'], reverse=True,
    )
    rows, used_matches = [], set()
    for item in ranked:
        if len(rows) >= 8:
            break
        number = str(item['赛事编号'])
        if number in used_matches:
            continue
        rows.append(item)
        used_matches.add(number)
    for row in rows:
        row.pop('_quality', None)
    return pd.DataFrame(rows, columns=columns)


def _combination_number(row: pd.Series, column: str) -> float:
    value = pd.to_numeric(row.get(column), errors='coerce')
    return float(value) if pd.notna(value) else float('nan')


def _match_identity(value) -> str:
    text = str(value or '').strip()
    if not text or text.lower() == 'nan':
        return ''
    try:
        return str(int(float(text)))
    except (TypeError, ValueError):
        return text


def _half_time_combination_evaluation(
        row: pd.Series,
        example_stake: float,
) -> dict | None:
    """Evaluate one three-way half/full cover without deciding whether to bet."""
    conflict = str(row.get('阵容方向冲突') or '').strip().lower()
    if conflict in {'true', '1', 'yes', '是'}:
        return None
    if str(row.get('阵容预警级别') or '').strip() == '高':
        return None
    directions = ('胜', '平', '负')
    formal = np.array([
        _combination_number(row, f'正式半场{direction}概率')
        for direction in directions
    ], dtype=float)
    monte = np.array([
        _combination_number(row, f'模拟半场{direction}概率')
        for direction in directions
    ], dtype=float)
    if (
            not np.isfinite(formal).all() or not np.isfinite(monte).all()
            or formal.sum() <= 0 or monte.sum() <= 0
    ):
        return None
    formal /= formal.sum()
    monte /= monte.sum()
    target_index = int(np.argmax(formal))
    if int(np.argmax(monte)) != target_index:
        return None

    group_odds = {}
    group_inverse = []
    for direction in directions:
        odds = np.array([
            _combination_number(row, f'官方半全场{label}奖金')
            for label in HALF_TIME_COMBINATION_GROUPS[direction]
        ], dtype=float)
        if not np.isfinite(odds).all() or np.any(odds <= 1.0):
            return None
        group_odds[direction] = odds
        group_inverse.append(float((1.0 / odds).sum()))
    market_half = np.asarray(group_inverse, dtype=float)
    market_half /= market_half.sum()
    if int(np.argmax(market_half)) != target_index:
        return None

    target = directions[target_index]
    labels = HALF_TIME_COMBINATION_GROUPS[target]
    odds = group_odds[target]
    inverse_sum = float((1.0 / odds).sum())
    combined_odds = 1.0 / inverse_sum
    break_even = inverse_sum
    formal_probability = float(formal[target_index])
    monte_probability = float(monte[target_index])
    formal_sorted = np.sort(formal)
    formal_margin = float(formal_sorted[-1] - formal_sorted[-2])
    agreement_gap = abs(formal_probability - monte_probability)
    conservative_probability = min(formal_probability, monte_probability) - 0.03
    model_edge = conservative_probability - break_even
    raw_ev = conservative_probability * combined_odds - 1.0
    formal_ev = formal_probability * combined_odds - 1.0
    monte_ev = monte_probability * combined_odds - 1.0
    relative_score = min(100.0, max(
        0.0,
        35.0
        + formal_probability * 30.0
        + monte_probability * 20.0
        + min(25.0, formal_margin * 150.0)
        + max(0.0, 10.0 - agreement_gap * 100.0),
    ))
    strict_score = min(100.0, max(
        0.0,
        50.0
        + min(25.0, max(0.0, model_edge) * 250.0)
        + min(15.0, formal_margin * 150.0)
        + max(0.0, 10.0 - agreement_gap * 100.0),
    ))
    model_source = str(row.get('半全场模型来源') or '').strip()
    verified = model_source.endswith('专用半全场模型（已验证）')
    strict_reasons = []
    if not verified:
        strict_reasons.append('缺独立验证')
    if formal_margin < 0.04:
        strict_reasons.append('方向领先不足4pp')
    if combined_odds < 1.45 or combined_odds > 2.20:
        strict_reasons.append('组合赔率超出1.45～2.20')
    if model_edge < 0.02:
        strict_reasons.append(f'低于正式门槛{model_edge:+.1%}')
    if formal_ev < 0.06 or monte_ev < 0.02:
        strict_reasons.append('模型收益验证不足')
    if strict_score < 72.0:
        strict_reasons.append('正式含金量不足72')

    weights = (1.0 / odds) / inverse_sum
    allocations = weights * float(example_stake)
    card_day = _ticket_card_date(row.get('比赛时间'), row.get('赛事编号'))
    day_text = (
        card_day.isoformat() if card_day is not None
        else str(row.get('比赛时间') or '')[:10]
    )
    return {
        '_strict': not strict_reasons,
        '_strict_reasons': strict_reasons,
        '_strict_score': strict_score,
        '_relative_score': relative_score,
        '_observation_rank': (
            formal_probability + monte_probability + formal_margin - agreement_gap
        ),
        '_raw_ev': raw_ev,
        '比赛日期': day_text,
        '比赛ID': _match_identity(row.get('比赛ID')),
        '赛事编号': row.get('赛事编号', ''),
        '联赛': row.get('联赛', ''),
        '对阵': f'{row.get("主队", "")} vs {row.get("客队", "")}',
        '目标半场': target,
        '组合玩法': ' / '.join(
            f'{label}@{odd:.2f}' for label, odd in zip(labels, odds)
        ),
        '组合赔率': round(combined_odds, 3),
        '半场含金量': f'{relative_score:.0f}/100',
        '正式半场概率': f'{formal_probability:.1%}',
        '蒙特半场概率': f'{monte_probability:.1%}',
        '市场半场概率': f'{float(market_half[target_index]):.1%}',
        '保本命中率': f'{break_even:.1%}',
        '模型优势': f'{model_edge:+.1%}｜EV {raw_ev:+.1%}',
        '示例本金': round(float(example_stake), 2),
        '等回报分配': ' / '.join(
            f'{label}¥{amount:.0f}' for label, amount in zip(labels, allocations)
        ),
        '策略状态': (
            '达到正式门槛，将进入真实账本'
            if not strict_reasons else '仅观察｜' + '；'.join(strict_reasons)
        ),
        '模型来源': model_source,
        '冻结时间': datetime.now().isoformat(timespec='seconds'),
    }


def build_half_time_combinations(
        predictions: pd.DataFrame,
        future_only: bool = True,
        example_stake: float = HALF_TIME_COMBINATION_STAKE,
) -> pd.DataFrame:
    """Build one auditable half-time dutching candidate per lottery-card day.

    Three half/full outcomes sharing the same half-time result are covered with
    inverse-odds stakes.  The full-time result is therefore irrelevant when the
    selected half-time direction lands.  A row is emitted only when the formal
    half/full model, official HAFU market and independent Monte Carlo all select
    the same half-time direction and both models clear the payout break-even.
    """
    columns = [
        '比赛日期', '比赛ID', '赛事编号', '联赛', '对阵', '目标半场',
        '组合玩法', '组合赔率', '半场含金量', '正式半场概率', '蒙特半场概率',
        '市场半场概率', '保本命中率', '模型优势', '示例本金',
        '等回报分配', '策略状态', '模型来源', '冻结时间',
    ]
    if predictions.empty:
        return pd.DataFrame(columns=columns)
    source = _upcoming_predictions(predictions) if future_only else predictions.copy()
    active = _sort_by_match_number(source).reset_index(drop=True)
    candidates_by_day: dict[str, list[dict]] = {}

    for _, row in active.iterrows():
        item = _half_time_combination_evaluation(row, example_stake)
        if item is None or not item['_strict']:
            continue
        item['半场含金量'] = f'{item["_strict_score"]:.0f}/100'
        item['策略状态'] = '三方同向候选｜高方差｜先记账验证'
        item['_quality'] = item['_strict_score'] + item['_raw_ev']
        candidates_by_day.setdefault(item['比赛日期'], []).append(item)

    rows = []
    for day in sorted(candidates_by_day):
        rows.append(max(candidates_by_day[day], key=lambda item: item['_quality']))
    for row in rows:
        for key in tuple(row):
            if key.startswith('_'):
                row.pop(key, None)
    return pd.DataFrame(rows, columns=columns)


def build_half_time_observations(
        predictions: pd.DataFrame,
        future_only: bool = True,
        example_stake: float = HALF_TIME_COMBINATION_STAKE,
) -> pd.DataFrame:
    """Show the strongest three-way-aligned row per card day without betting it."""
    columns = [
        '比赛日期', '赛事编号', '联赛', '对阵', '目标半场', '组合玩法',
        '组合赔率', '相对含金量', '正式半场概率', '蒙特半场概率',
        '市场半场概率', '保本命中率', '模型优势', '观察结论', '模型来源',
    ]
    if predictions.empty:
        return pd.DataFrame(columns=columns)
    source = _upcoming_predictions(predictions) if future_only else predictions.copy()
    active = _sort_by_match_number(source).reset_index(drop=True)
    observations_by_day: dict[str, list[dict]] = {}
    for _, row in active.iterrows():
        item = _half_time_combination_evaluation(row, example_stake)
        if item is None:
            continue
        observations_by_day.setdefault(item['比赛日期'], []).append(item)

    rows = []
    for day in sorted(observations_by_day):
        item = max(
            observations_by_day[day],
            key=lambda candidate: candidate['_observation_rank'],
        )
        rows.append({
            '比赛日期': item['比赛日期'],
            '赛事编号': item['赛事编号'],
            '联赛': item['联赛'],
            '对阵': item['对阵'],
            '目标半场': item['目标半场'],
            '组合玩法': item['组合玩法'],
            '组合赔率': item['组合赔率'],
            '相对含金量': item['半场含金量'],
            '正式半场概率': item['正式半场概率'],
            '蒙特半场概率': item['蒙特半场概率'],
            '市场半场概率': item['市场半场概率'],
            '保本命中率': item['保本命中率'],
            '模型优势': item['模型优势'],
            '观察结论': item['策略状态'],
            '模型来源': item['模型来源'],
        })
    return pd.DataFrame(rows, columns=columns)


def _load_half_time_combination_snapshot(day: str) -> pd.DataFrame:
    path = HALF_TIME_COMBINATION_ROOT / f'{day}.csv'
    if not path.exists():
        return pd.DataFrame()
    try:
        return pd.read_csv(path)
    except (OSError, pd.errors.EmptyDataError, UnicodeError):
        return pd.DataFrame()


def _save_half_time_combination_snapshot(frame: pd.DataFrame) -> None:
    """Freeze first-seen candidates so later odds cannot rewrite the ledger."""
    if frame.empty or '比赛日期' not in frame.columns:
        return
    HALF_TIME_COMBINATION_ROOT.mkdir(parents=True, exist_ok=True)
    for day, rows in frame.groupby('比赛日期', sort=False):
        day_text = str(day or '').strip()
        if not re.fullmatch(r'\d{4}-\d{2}-\d{2}', day_text):
            continue
        existing = _load_half_time_combination_snapshot(day_text)
        combined = rows.copy().reset_index(drop=True)
        if not existing.empty:
            existing_keys = set(
                existing.get('比赛ID', pd.Series(dtype=str)).map(_match_identity)
            )
            existing_numbers = set(
                existing.get('赛事编号', pd.Series(dtype=str)).astype(str)
            )
            unseen = combined.loc[
                ~combined.get('比赛ID', pd.Series('', index=combined.index)).map(
                    _match_identity,
                ).isin(existing_keys)
                & ~combined.get('赛事编号', pd.Series('', index=combined.index)).astype(
                    str,
                ).isin(existing_numbers)
            ]
            combined = pd.concat([existing, unseen], ignore_index=True, sort=False)
        path = HALF_TIME_COMBINATION_ROOT / f'{day_text}.csv'
        temporary = path.with_suffix('.tmp')
        combined.to_csv(temporary, index=False, encoding='utf-8-sig')
        temporary.replace(path)


def build_half_time_combination_ledger(
        example_stake: float = HALF_TIME_COMBINATION_STAKE,
) -> tuple[pd.DataFrame, dict]:
    """Settle every frozen combination without deleting losses or postponements."""
    columns = [
        '日期', '赛事编号', '对阵', '组合玩法', '目标半场', '半场含金量',
        '组合赔率', '金额', '半场赛果', '本次盈利', '累计盈利', '结算状态',
    ]
    snapshots = []
    if HALF_TIME_COMBINATION_ROOT.exists():
        for path in sorted(HALF_TIME_COMBINATION_ROOT.glob('*.csv')):
            try:
                frame = pd.read_csv(path)
            except (OSError, pd.errors.EmptyDataError, UnicodeError):
                continue
            if not frame.empty:
                snapshots.append(frame)
    if not snapshots:
        return pd.DataFrame(columns=columns), {
            'total': 0, 'settled': 0, 'pending': 0, 'hits': 0,
            'stake': 0.0, 'settled_stake': 0.0, 'profit': 0.0,
            'roi': None, 'hit_rate': None,
        }
    frozen = pd.concat(snapshots, ignore_index=True, sort=False)
    frozen = frozen.drop_duplicates(['比赛日期', '赛事编号'], keep='first')
    settled = pd.DataFrame()
    if SETTLED_PREDICTIONS_PATH.exists():
        try:
            settled = pd.read_csv(SETTLED_PREDICTIONS_PATH)
        except (OSError, pd.errors.EmptyDataError, UnicodeError):
            settled = pd.DataFrame()
    by_id = {}
    by_date_number = {}
    if not settled.empty:
        for _, result in settled.iterrows():
            match_id = _match_identity(result.get('match_id'))
            if match_id:
                by_id[match_id] = result
            number = str(result.get('match_number') or '').strip()
            match_day = str(result.get('match_date') or '')[:10]
            if number and re.fullmatch(r'\d{4}-\d{2}-\d{2}', match_day):
                by_date_number[(match_day, number)] = result

    rows = []
    cumulative = 0.0
    settled_count = hits = 0
    for _, item in frozen.sort_values(['比赛日期', '赛事编号']).iterrows():
        match_id = _match_identity(item.get('比赛ID'))
        number = str(item.get('赛事编号') or '').strip()
        result = by_id.get(match_id) if match_id else None
        if result is None:
            result = by_date_number.get((str(item.get('比赛日期') or ''), number))
        actual_half_full = str(
            result.get('actual_half_full') if result is not None else ''
        ).strip()
        if actual_half_full.lower() == 'nan':
            actual_half_full = ''
        target = str(item.get('目标半场') or '').strip()
        combined_odds = _combination_number(item, '组合赔率')
        amount = _combination_number(item, '示例本金')
        if not np.isfinite(amount) or amount <= 0:
            amount = float(example_stake)
        if actual_half_full[:1] and actual_half_full[:1] in '胜平负':
            hit = actual_half_full[:1] == target
            profit = amount * (combined_odds - 1.0) if hit else -amount
            cumulative += profit
            settled_count += 1
            hits += int(hit)
            status = '✓ 命中' if hit else '✕ 未中'
            actual_text = f'半场{actual_half_full[:1]}'
            profit_text = f'{profit:+.0f}'
            cumulative_text = f'{cumulative:+.0f}'
        else:
            status = '○ 延期/待定'
            actual_text = '待官方赛果'
            profit_text = '待定'
            cumulative_text = f'{cumulative:+.0f}'
        rows.append({
            '日期': item.get('比赛日期', ''),
            '赛事编号': number,
            '对阵': item.get('对阵', ''),
            '组合玩法': item.get('组合玩法', ''),
            '目标半场': target,
            '半场含金量': item.get('半场含金量', ''),
            '组合赔率': combined_odds,
            '金额': amount,
            '半场赛果': actual_text,
            '本次盈利': profit_text,
            '累计盈利': cumulative_text,
            '结算状态': status,
        })
    total = len(rows)
    settled_stake = settled_count * float(example_stake)
    summary = {
        'total': total,
        'settled': settled_count,
        'pending': total - settled_count,
        'hits': hits,
        'stake': total * float(example_stake),
        'settled_stake': settled_stake,
        'profit': cumulative,
        'roi': cumulative / settled_stake if settled_stake else None,
        'hit_rate': hits / settled_count if settled_count else None,
    }
    return pd.DataFrame(rows, columns=columns), summary


def _save_daily_recommendation_snapshot(frame: pd.DataFrame) -> None:
    """Persist the recommendations that were actually shown to the user."""
    if frame.empty or '比赛日期' not in frame.columns:
        return
    DAILY_RECOMMENDATION_ROOT.mkdir(parents=True, exist_ok=True)
    for day, rows in frame.groupby('比赛日期', sort=False):
        day_text = str(day or '').strip()
        if not re.fullmatch(r'\d{4}-\d{2}-\d{2}', day_text):
            continue
        path = DAILY_RECOMMENDATION_ROOT / f'{day_text}.csv'
        combined = rows.copy().reset_index(drop=True)
        preserving_past_card = day_text < date.today().isoformat()
        if path.exists() and preserving_past_card:
            # Lottery cards legitimately continue after midnight.  Preserve
            # every recommendation already shown on the previous card and only
            # append fixtures that were never frozen.  Historical cards may
            # legitimately contain multiple legacy plays for one match; never
            # collapse or reinterpret those rows with today's selection rules.
            # Once no row from the old card is visible, this function receives
            # no group for that day and the audit file becomes immutable.
            existing = _load_daily_recommendation_snapshot(day_text)
            if not existing.empty and '赛事编号' in combined.columns:
                frozen_numbers = set(existing['赛事编号'].astype(str))
                unseen = combined.loc[
                    ~combined['赛事编号'].astype(str).isin(frozen_numbers)
                ]
                combined = pd.concat([existing, unseen], ignore_index=True, sort=False)
        # During the active card, persist exactly the latest list shown.  One
        # fixture owns one core play; a direction/play change replaces the old
        # row instead of inflating yesterday's review.
        keys = (
            [column for column in ('赛事编号', '推荐玩法') if column in combined.columns]
            if preserving_past_card else
            (['赛事编号'] if '赛事编号' in combined.columns else [])
        )
        if keys:
            combined = combined.drop_duplicates(keys, keep='last')
        temporary = path.with_suffix('.tmp')
        combined.to_csv(temporary, index=False, encoding='utf-8-sig')
        temporary.replace(path)


def _load_daily_recommendation_snapshot(day: str) -> pd.DataFrame:
    path = DAILY_RECOMMENDATION_ROOT / f'{day}.csv'
    if not path.exists():
        return pd.DataFrame()
    try:
        return pd.read_csv(path)
    except (OSError, pd.errors.EmptyDataError, UnicodeError):
        return pd.DataFrame()


def build_yesterday_recommendation_review() -> tuple[pd.DataFrame, str]:
    """Score the frozen recommendations that were actually shown yesterday."""
    columns = [
        '赛事编号', '联赛', '对阵', '推荐玩法', '昨日推荐',
        '推荐等级', '正式模型概率', '价值评估', '蒙特卡洛是否同向',
        '完场比分', '复盘结果',
        '失败原因', '命中状态',
    ]
    details, summary = load_yesterday_hit_report()
    review_date = str(summary.get('date') or '')
    if not review_date:
        return pd.DataFrame(columns=columns), review_date
    recommendations = _load_daily_recommendation_snapshot(review_date)
    if recommendations.empty:
        # Never apply today's thresholds retroactively.  Without a frozen
        # snapshot there is no honest way to know what the user saw yesterday.
        return pd.DataFrame(columns=columns), review_date
    detail_by_number = (
        details.drop_duplicates('赛事编号', keep='last').set_index('赛事编号')
        if not details.empty else pd.DataFrame()
    )
    result_columns = {
        '胜负': '胜负', '胜平负': '胜负', '胜平负·平局': '胜负',
        '让球': '让球（首/次）', '让球胜平负': '让球（首/次）', '大小球': '大小球',
        '半全场': '半全场（首/次）', '比分': '比分（首/次1/次2/冷/进）',
    }
    rows = []
    for _, recommendation in recommendations.iterrows():
        number = recommendation['赛事编号']
        if number not in detail_by_number.index:
            rows.append({
                '赛事编号': number,
                '联赛': recommendation.get('联赛', ''),
                '对阵': recommendation.get('对阵', ''),
                '推荐玩法': recommendation.get('推荐玩法', ''),
                '昨日推荐': recommendation.get('重点选项', ''),
                '推荐等级': recommendation.get('推荐等级', ''),
                '正式模型概率': recommendation.get(
                    '正式模型概率', recommendation.get('模型概率', ''),
                ),
                '价值评估': recommendation.get('价值评估', ''),
                '蒙特卡洛是否同向': recommendation.get('蒙特卡洛是否同向', ''),
                '完场比分': '',
                '复盘结果': '官方赛果未补齐（延期或未完场）',
                '失败原因': '待官方赛果，不计失败',
                '命中状态': '○ 待复盘',
            })
            continue
        detail = detail_by_number.loc[number]
        market = recommendation['推荐玩法']
        result = str(detail.get(result_columns.get(market, '')) or '')
        selected = str(recommendation['重点选项'] or '').replace('★', '').strip()
        if market == '总进球':
            actual_score = str(detail.get('完场比分') or '').strip()
            score_match = re.fullmatch(r'(\d+)-(\d+)', actual_score)
            actual_total = sum(map(int, score_match.groups())) if score_match else None
            actual = '7+球' if actual_total is not None and actual_total >= 7 else (
                f'{actual_total}球' if actual_total is not None else ''
            )
            hit = bool(selected and actual and selected == actual)
            result = f'{selected or "—"} → {actual or "—"}（{"命中" if hit else "未中"}）'
        elif market == '比分':
            actual = str(detail.get('完场比分') or '').strip()
            hit = bool(selected and actual and selected == actual)
            result = f'{selected or "—"} → {actual or "—"}（{"命中" if hit else "未中"}）'
        elif market == '半全场':
            actual_match = re.search(r'→\s*([胜平负]{2})', result)
            actual = actual_match.group(1) if actual_match else ''
            hit = bool(selected and actual and selected == actual)
            result = f'{selected or "—"} → {actual or "—"}（{"命中" if hit else "未中"}）'
        elif market in ('让球', '让球胜平负'):
            actual_match = re.search(r'→\s*让([胜平负])', result)
            actual = actual_match.group(1) if actual_match else ''
            selected_pick = selected[-1:] if selected[-1:] in '胜平负' else ''
            hit = bool(selected_pick and actual and selected_pick == actual)
            result = (
                f'{selected or "—"} → 让{actual}（{"命中" if hit else "未中"}）'
                if actual else '未开盘'
            )
        elif market == '胜平负·平局':
            hit = '→ 平（命中）' in result
        else:
            hit = '（命中）' in result
        failure_reason = ''
        if not hit:
            market_note = str(recommendation.get('盘口验证') or '')
            if '反向' in market_note or '走弱' in market_note:
                failure_reason = '盘口临场反转'
            elif market in ('比分', '半全场', '总进球'):
                failure_reason = '主方向或具体玩法选择偏差'
            else:
                failure_reason = '正式模型方向错误'
        rows.append({
            '赛事编号': number,
            '联赛': recommendation['联赛'],
            '对阵': recommendation['对阵'],
            '推荐玩法': market,
            '昨日推荐': recommendation['重点选项'],
            '推荐等级': recommendation.get('推荐等级', ''),
            '正式模型概率': recommendation.get(
                '正式模型概率', recommendation.get('模型概率', ''),
            ),
            '价值评估': recommendation.get('价值评估', ''),
            '蒙特卡洛是否同向': recommendation.get('蒙特卡洛是否同向', ''),
            '完场比分': detail.get('完场比分', ''),
            '复盘结果': result,
            '失败原因': failure_reason,
            '命中状态': '✓ 命中' if hit else '✕ 未中',
        })
    return pd.DataFrame(rows, columns=columns), review_date


class YesterdayRecommendationReviewDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        frame, review_date = build_yesterday_recommendation_review()
        self.setWindowTitle(f'{review_date or "昨日"} 每日推荐复盘')
        self.resize(1180, 560)
        root = QVBoxLayout(self)
        if frame.empty:
            root.addWidget(QLabel(
                '昨日重点推荐快照未保存，或官方赛果尚未补齐。\n'
                '为保证准确，本页不会使用今天的门槛倒推昨日推荐。'
            ))
            return
        settled = frame['命中状态'].ne('○ 待复盘')
        settled_count = int(settled.sum())
        pending_count = int((~settled).sum())
        hits = int(frame.loc[settled, '命中状态'].eq('✓ 命中').sum())
        breakdown = []
        for market, part in frame.loc[settled].groupby('推荐玩法', sort=False):
            market_hits = int(part['命中状态'].eq('✓ 命中').sum())
            breakdown.append(f'{market}{market_hits}/{len(part)}')
        breakdown_text = '；'.join(breakdown)
        root.addWidget(QLabel(
            (
                f'{review_date} 每日推荐：已结算 {hits}/{settled_count} 命中，'
                f'待复盘 {pending_count} 项（分玩法：{breakdown_text or "暂无"}）'
            ) if settled_count else (
                f'{review_date} 每日推荐：{pending_count} 项均为延期或未完场，待复盘'
            )
        ))
        table = ExcelTable(self, frame, readonly=True, supports_sorting=True)
        status_column = frame.columns.get_loc('命中状态')
        for row in range(table.rowCount()):
            item = table.item(row, status_column)
            hit = item.text().startswith('✓')
            pending = item.text().startswith('○')
            foreground = '#9a6700' if pending else ('#137333' if hit else '#c62828')
            background = '#fff8df' if pending else ('#eef8f0' if hit else '#fff1f1')
            item.setForeground(QBrush(QColor(foreground)))
            item.setBackground(QBrush(QColor(background)))
            font = QFont(item.font())
            font.setBold(True)
            item.setFont(font)
        root.addWidget(table)


class HalfTimeCombinationLedgerDialog(QDialog):
    """Frozen, no-cherry-picking audit of half-time dutching candidates."""

    def __init__(self, predictions: pd.DataFrame, parent=None):
        super().__init__(parent)
        self._predictions = predictions.copy()
        self.setWindowTitle('半场组合策略与真实盈亏账本')
        self.setWindowFlags(
            Qt.WindowType.Window | Qt.WindowType.WindowMinimizeButtonHint
            | Qt.WindowType.WindowMaximizeButtonHint | Qt.WindowType.WindowCloseButtonHint
        )
        self.resize(1500, 760)
        root = QVBoxLayout(self)
        observations = build_half_time_observations(predictions)
        candidates = build_half_time_combinations(predictions)
        _save_half_time_combination_snapshot(candidates)
        ledger, summary = build_half_time_combination_ledger()
        observation_title = QLabel(
            '今日相对最优观察（只比较方向稳定性，不计入命中率与ROI）'
        )
        observation_title.setStyleSheet(
            'color:#8a4b08;background:#fff8df;border:1px solid #eed28a;'
            'padding:5px;font-size:13px;font-weight:600;'
        )
        root.addWidget(observation_title)
        if observations.empty:
            root.addWidget(QLabel(
                '当前没有正式模型、蒙特和官方半全场盘口三方同向的观察项。'
            ))
        else:
            observation_table = ExcelTable(
                self, observations, readonly=True, supports_sorting=True,
            )
            conclusion_column = observations.columns.get_loc('观察结论')
            for row in range(observation_table.rowCount()):
                item = observation_table.item(row, conclusion_column)
                if item is None:
                    continue
                formal = item.text().startswith('达到正式门槛')
                item.setForeground(QBrush(QColor('#137333' if formal else '#9a6700')))
                item.setBackground(QBrush(QColor('#eef8f0' if formal else '#fff8df')))
                font = QFont(item.font())
                font.setBold(True)
                item.setFont(font)
            root.addWidget(observation_table)

        ledger_title = QLabel('正式组合推荐与冻结真实盈亏账本')
        ledger_title.setStyleSheet('font-size:13px;font-weight:600;padding-top:4px;')
        root.addWidget(ledger_title)
        roi = summary.get('roi')
        hit_rate = summary.get('hit_rate')
        summary_text = (
            f'组合总数 {summary["total"]}｜已结算 {summary["settled"]}｜'
            f'待定 {summary["pending"]}｜演示总投入 ¥{summary["stake"]:,.0f}｜'
            f'已结算投入 ¥{summary["settled_stake"]:,.0f}｜'
            f'累计盈亏 {summary["profit"]:+,.0f}｜'
            f'ROI {roi:.1%}｜命中率 {hit_rate:.1%}'
            if roi is not None and hit_rate is not None else (
                f'当前冻结候选 {summary["total"]} 场，尚无已结算样本。'
            )
        )
        headline = QLabel(summary_text)
        headline.setWordWrap(True)
        headline.setStyleSheet(
            'color:#12344d;background:#eef6fb;border:1px solid #bdd8e8;'
            'padding:6px;font-size:13px;font-weight:600;'
        )
        root.addWidget(headline)
        note = QLabel(
            '组合规则：同一半场方向的三项半全场用反赔率分配，命中目标半场即可获得近似等额回报。'
            '仅收录正式专用模型、官方半全场市场和独立蒙特同向，且模型概率高于保本线的场次；'
            '“半场含金量”综合模型优势、方向领先幅度和蒙特一致度，只是筛选分而非命中概率；'
            '上方“相对最优观察”只比较当日方向稳定性，即使未过收益线也显示，但绝不进入正式账本；'
            '首次出现即冻结，输单不删除，延期不算输。金额为每场1000元演示账本，不代表投注指令。'
        )
        note.setWordWrap(True)
        root.addWidget(note)
        if ledger.empty:
            root.addWidget(QLabel('当前没有通过组合门控的场次，继续等待，不为凑数强推。'))
            return
        table = ExcelTable(self, ledger, readonly=True, supports_sorting=True)
        status_column = ledger.columns.get_loc('结算状态')
        profit_column = ledger.columns.get_loc('本次盈利')
        for row in range(table.rowCount()):
            status_item = table.item(row, status_column)
            profit_item = table.item(row, profit_column)
            status = status_item.text() if status_item is not None else ''
            hit = status.startswith('✓')
            pending = status.startswith('○')
            foreground = '#9a6700' if pending else ('#137333' if hit else '#c62828')
            background = '#fff8df' if pending else ('#eef8f0' if hit else '#fff1f1')
            for item in (status_item, profit_item):
                if item is None:
                    continue
                item.setForeground(QBrush(QColor(foreground)))
                item.setBackground(QBrush(QColor(background)))
                font = QFont(item.font())
                font.setBold(True)
                item.setFont(font)
        root.addWidget(table)


class DailyRecommendationsDialog(QDialog):
    """Focused list of the strongest daily options across all markets."""

    def __init__(self, predictions: pd.DataFrame, parent=None):
        super().__init__(parent)
        self._predictions = predictions.copy()
        self.setWindowTitle('每日重点推荐')
        self.setWindowFlags(
            Qt.WindowType.Window | Qt.WindowType.WindowMinimizeButtonHint
            | Qt.WindowType.WindowMaximizeButtonHint | Qt.WindowType.WindowCloseButtonHint
        )
        self.resize(1440, 620)
        root = QVBoxLayout(self)
        header = QHBoxLayout()
        notice = QLabel(
            '每天目标6～8场、正期望优先；核心重点与可买优选分级展示，'
            '三方同向但价值不足、或蒙特反向时灰色显示观察且建议不投注；'
            '◎最佳比分和◆高倍候选为醒目参考，高倍项不等于重点；'
            '比分Top3和半全场前两项仅供参考；半场组合另设冻结盈亏账本。'
        )
        notice.setWordWrap(True)
        header.addWidget(notice, 1)
        review_button = QPushButton('昨日推荐复盘')
        review_button.clicked.connect(self._open_yesterday_review)
        header.addWidget(review_button)
        combination_button = QPushButton('半场组合账本')
        combination_button.clicked.connect(self._open_half_time_combinations)
        header.addWidget(combination_button)
        root.addLayout(header)
        self._review_dialog = None
        self._combination_dialog = None
        frame = build_daily_recommendations(predictions)
        _save_daily_recommendation_snapshot(frame)
        if frame.empty:
            root.addWidget(QLabel('当前没有达到推荐门槛的选项。'))
            return
        table = ExcelTable(self, frame, readonly=True, supports_sorting=True)
        if '重点选项' in frame.columns:
            column = frame.columns.get_loc('重点选项')
            grade_column = (
                frame.columns.get_loc('推荐等级')
                if '推荐等级' in frame.columns else None
            )
            for row in range(table.rowCount()):
                item = table.item(row, column)
                grade = (
                    table.item(row, grade_column).text()
                    if grade_column is not None else '核心重点'
                )
                core = grade == '核心重点'
                candidate = grade == '可买优选'
                foreground = '#c62828' if core else '#9a6700' if candidate else '#5f6368'
                background = '#fff1f1' if core else '#fff8df' if candidate else '#f3f4f6'
                item.setForeground(QBrush(QColor(foreground)))
                item.setBackground(QBrush(QColor(background)))
                font = QFont(item.font())
                font.setBold(True)
                item.setFont(font)
        for column_name, foreground, background in (
                ('最佳比分', '#1565c0', '#eef5ff'),
                ('高倍候选', '#7b1fa2', '#f7efff')):
            if column_name not in frame.columns:
                continue
            column = frame.columns.get_loc(column_name)
            for row in range(table.rowCount()):
                item = table.item(row, column)
                if item is None or item.text().strip() in ('', '—'):
                    continue
                item.setForeground(QBrush(QColor(foreground)))
                item.setBackground(QBrush(QColor(background)))
                font = QFont(item.font())
                font.setBold(True)
                item.setFont(font)
        root.addWidget(table)

    def _open_yesterday_review(self):
        if self._review_dialog is not None:
            self._review_dialog.close()
            self._review_dialog.deleteLater()
        self._review_dialog = YesterdayRecommendationReviewDialog(self)
        self._review_dialog.show()
        self._review_dialog.raise_()
        self._review_dialog.activateWindow()

    def _open_half_time_combinations(self):
        if self._combination_dialog is not None:
            self._combination_dialog.close()
            self._combination_dialog.deleteLater()
        self._combination_dialog = HalfTimeCombinationLedgerDialog(
            self._predictions, self,
        )
        self._combination_dialog.show()
        self._combination_dialog.raise_()
        self._combination_dialog.activateWindow()


def filter_predictions_by_model(df: pd.DataFrame, model_key: str) -> pd.DataFrame:
    """Strictly isolate one dedicated league or the generic/market rows."""
    if df.empty or model_key in ('', ALL_MODELS):
        return df.copy()
    if model_key == SIMULATION_MODELS:
        simulations = pd.to_numeric(
            df.get('模拟次数', pd.Series(0, index=df.index)), errors='coerce',
        ).fillna(0)
        result = df.get(
            '模拟胜负', pd.Series('', index=df.index),
        ).fillna('').astype(str).str.strip()
        return df.loc[simulations.gt(0) | result.ne('')].copy()
    scopes = _prediction_model_scopes(df)
    if model_key == DEDICATED_MODELS:
        mask = scopes.ne('')
    elif model_key == GENERIC_MODELS:
        mask = scopes.eq('')
    else:
        mask = scopes.eq(model_key)
    return df.loc[mask].reset_index(drop=True)


class SportteryPredictionsDialog(QDialog):
    """Synchronize and display today's official Sporttery model predictions."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle('今日竞彩概率分析')
        # Force a regular native top-level window. On macOS a parented QDialog
        # can otherwise be presented as a sheet without usable title-bar
        # controls, leaving no close button or area from which to drag it.
        self.setWindowFlags(
            Qt.WindowType.Window
            | Qt.WindowType.WindowTitleHint
            | Qt.WindowType.WindowMinimizeButtonHint
            | Qt.WindowType.WindowMaximizeButtonHint
            | Qt.WindowType.WindowCloseButtonHint
        )
        self.resize(1250, 650)
        self._predictions = pd.DataFrame()
        self._table_container = QWidget()
        self._table_layout = QVBoxLayout(self._table_container)
        self._table_layout.setContentsMargins(0, 0, 0, 0)
        self._summary = QLabel()
        self._model_selector = QComboBox()
        self._date_selector = QComboBox()
        self._advice_selector = QComboBox()
        self._yesterday_dialog = None
        self._market_dialog = None
        self._daily_recommendations_dialog = None
        self._table = None
        self._odds_pulse_running = False
        self._odds_pulse_queue = Queue()
        self._lineup_running = False
        self._foreground_sync_running = False
        self._lineup_queue = Queue()
        self._trained_leagues = trained_dedicated_leagues()
        self._build_ui()
        self._load_saved_reports()
        self._odds_pulse_timer = QTimer(self)
        self._odds_pulse_timer.setInterval(5 * 60_000)
        self._odds_pulse_timer.timeout.connect(self._start_odds_pulse)
        self._odds_pulse_timer.start()
        self._odds_pulse_result_timer = QTimer(self)
        self._odds_pulse_result_timer.setInterval(300)
        self._odds_pulse_result_timer.timeout.connect(self._collect_odds_pulse)
        self._odds_pulse_result_timer.start()
        self._lineup_timer = QTimer(self)
        self._lineup_timer.setSingleShot(True)
        self._lineup_timer.timeout.connect(self._start_lineup_supervision)
        self._lineup_result_timer = QTimer(self)
        self._lineup_result_timer.setInterval(500)
        self._lineup_result_timer.timeout.connect(self._collect_lineup_supervision)
        QTimer.singleShot(12_000, self._start_odds_pulse)
        QTimer.singleShot(15_000, self._schedule_lineup_supervision)

    def _schedule_lineup_supervision(self):
        """Poll only pending fixtures inside the 90-minute lineup window."""
        self._lineup_timer.stop()
        if (
                not lineup_api_configured() or self._predictions.empty
                or '比赛时间' not in self._predictions.columns
        ):
            return
        kickoff = pd.to_datetime(
            self._predictions.get('比赛时间'), errors='coerce',
        )
        now = pd.Timestamp.now()
        minutes = (kickoff - now).dt.total_seconds() / 60.0
        status = self._predictions.get(
            '首发状态', pd.Series('', index=self._predictions.index),
        ).fillna('').astype(str)
        pending = minutes.between(-15, 90) & status.ne('已确认')
        if not pending.any():
            return
        interval = lineup_poll_interval_seconds(float(minutes.loc[pending].min()))
        self._lineup_timer.start(max(1_000, int(interval * 1000)))
        self._summary.setToolTip(
            f'阵容自动督导已开启：{int(pending.sum())}场待确认，'
            f'{interval // 60}分钟后复查。'
        )

    def _start_lineup_supervision(self):
        if self._lineup_running or self._foreground_sync_running:
            return
        self._lineup_running = True
        self._lineup_result_timer.start()
        Thread(target=self._fetch_lineup_supervision, daemon=True).start()

    def _fetch_lineup_supervision(self):
        try:
            self._lineup_queue.put(('ok', run_daily_sporttery()))
        except Exception as error:
            self._lineup_queue.put(('error', str(error)))

    def _collect_lineup_supervision(self):
        try:
            state, payload = self._lineup_queue.get_nowait()
        except Empty:
            return
        self._lineup_running = False
        self._lineup_result_timer.stop()
        if state == 'ok' and payload is not None:
            self._predictions, _ = payload
            _save_daily_recommendation_snapshot(
                build_daily_recommendations(self._predictions),
            )
            self._refresh_model_selector()
            self._refresh_date_selector()
            self._refresh_advice_selector()
            self._render()
        self._schedule_lineup_supervision()

    def _start_odds_pulse(self):
        """Persist actual market changes while the prediction window is open."""
        if self._odds_pulse_running:
            return
        self._odds_pulse_running = True
        Thread(target=self._fetch_odds_pulse, daemon=True).start()

    def _fetch_odds_pulse(self):
        try:
            matches = SportteryMobileClient(timeout=8.0, retries=1).selling_matches()
            saved = record_odds_snapshots(matches)
            self._odds_pulse_queue.put(('ok', len(matches), saved))
        except Exception as error:
            self._odds_pulse_queue.put(('error', 0, str(error)))

    def _collect_odds_pulse(self):
        try:
            state, count, payload = self._odds_pulse_queue.get_nowait()
        except Empty:
            return
        self._odds_pulse_running = False
        if state == 'ok':
            self._summary.setToolTip(
                f'后台盘口采集：{count}场，新增真实变化 {payload} 条。'
            )

    def closeEvent(self, event):
        self._odds_pulse_timer.stop()
        self._odds_pulse_result_timer.stop()
        self._lineup_timer.stop()
        self._lineup_result_timer.stop()
        super().closeEvent(event)

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(6, 6, 6, 6)
        root.setSpacing(5)

        root.addWidget(_WindowDragBar(self.windowTitle(), self))

        action_bar = QHBoxLayout()
        action_bar.setSpacing(5)
        sync_button = QPushButton('同步赔率并更新分析')
        sync_button.setToolTip('重新获取官方最新赔率；建议在安全截止前20～60分钟执行')
        sync_button.clicked.connect(self._sync)
        review_button = QPushButton('补赛果并复盘')
        review_button.clicked.connect(self._review_history)
        yesterday_button = QPushButton('昨日命中复盘')
        yesterday_button.setToolTip('查看昨日已结算场次的逐项命中明细和简短规律')
        yesterday_button.clicked.connect(self._show_yesterday_hits)
        market_button = QPushButton('市场走势')
        market_button.setObjectName('marketTrendButton')
        market_button.setToolTip('查看欧赔概率、让球和大小球的初盘到临盘走势')
        market_button.clicked.connect(self._show_market_trends)
        daily_button = QPushButton('每日推荐')
        daily_button.setToolTip('集中查看胜负、让球、大小球、比分和半全场重点选项')
        daily_button.clicked.connect(self._show_daily_recommendations)
        export_button = QPushButton('导出 Excel')
        export_button.clicked.connect(self._export)
        for button, width in (
                (sync_button, 136), (review_button, 102),
                (yesterday_button, 100), (market_button, 80), (daily_button, 80),
                (export_button, 78)):
            button.setFixedSize(width, 25)
            action_bar.addWidget(button)
        action_bar.addStretch(1)
        self._summary.setObjectName('summaryLabel')
        self._summary.setMinimumWidth(195)
        action_bar.addWidget(self._summary)
        root.addLayout(action_bar)

        filter_panel = QWidget(self)
        filter_panel.setObjectName('filterPanel')
        filters = QHBoxLayout(filter_panel)
        filters.setContentsMargins(4, 3, 4, 3)
        filters.setSpacing(5)
        filters.addWidget(QLabel('模型'))
        self._model_selector.setMinimumWidth(220)
        self._model_selector.setFixedHeight(24)
        self._model_selector.currentIndexChanged.connect(lambda _: self._render())
        filters.addWidget(self._model_selector)
        filters.addSpacing(6)
        filters.addWidget(QLabel('比赛日期'))
        self._date_selector.setMinimumWidth(165)
        self._date_selector.setFixedHeight(24)
        self._date_selector.currentIndexChanged.connect(lambda _: self._render())
        filters.addWidget(self._date_selector)
        filters.addStretch(1)
        root.addWidget(filter_panel)
        risk_notice = QLabel(
            '概率不是赛果保证；优先查看样本量、覆盖率、阵容状态与模型分歧。'
        )
        risk_notice.setObjectName('riskNotice')
        risk_notice.setToolTip(
            '比分和半全场属于高方差市场；历史命中率只描述指定样本与时间窗口。'
        )
        root.addWidget(risk_notice)
        root.addWidget(self._table_container)

        self.setStyleSheet('''
            QWidget#windowDragBar {
                background: #e8edf1; border: 1px solid #b8c0c7;
                border-radius: 3px;
            }
            QLabel#windowDragTitle {
                color: #202020; font-size: 13px; font-weight: 600;
            }
            QPushButton#windowCloseButton {
                color: #303030; background: transparent;
                border: 0; font-size: 16px; font-weight: 600; padding: 0;
            }
            QPushButton#windowCloseButton:hover {
                color: #ffffff; background: #d93025; border-radius: 3px;
            }
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #ffffff, stop:0.55 #f5f6f7, stop:1 #e2e5e8);
                color: #202020; border: 1px solid #aeb5bb;
                border-bottom: 2px solid #8f979e; border-radius: 3px;
                padding: 1px 6px 2px 6px; font-size: 12px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #ffffff, stop:1 #dceaf7);
                border-color: #7199bd; border-bottom-color: #557d9f;
            }
            QPushButton:pressed {
                background: #d9e4ee; border: 1px solid #7f8b94;
                padding-top: 2px; padding-bottom: 1px;
            }
            QWidget#filterPanel {
                border-top: 1px solid #d7d7d7; border-bottom: 1px solid #d7d7d7;
            }
            QComboBox {
                color: #202020;
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #ffffff, stop:1 #e7eaed);
                border: 1px solid #aeb5bb; border-bottom: 2px solid #90989f;
                border-radius: 3px; padding: 0 22px 0 5px; font-size: 12px;
            }
            QComboBox:hover { border-color: #6f9dcc; }
            QComboBox::drop-down {
                width: 18px; border-left: 1px solid #b8bec4;
                background: #e8ebee;
            }
            QComboBox QAbstractItemView {
                color: #202020; background: #ffffff;
                border: 1px solid #90989f; outline: 0;
                selection-background-color: #2f79bd;
                selection-color: #ffffff;
                font-size: 12px;
            }
            QLabel#summaryLabel {
                color: #404040; padding: 3px 5px; font-size: 12px;
            }
            QLabel#riskNotice {
                color: #7a4b00; background: #fff8e6;
                border: 1px solid #ead7a5; border-radius: 3px;
                padding: 4px 7px; font-size: 12px;
            }
            QTableWidget {
                color: #202020; background: #ffffff;
                alternate-background-color: #f6f8fa;
                gridline-color: #cfd4d8; font-size: 12px;
            }
            QHeaderView::section {
                color: #202020;
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #ffffff, stop:1 #dfe3e6);
                border: 0; border-right: 1px solid #b8bec3;
                border-bottom: 1px solid #969da3;
                padding: 2px 4px; font-size: 12px; font-weight: 600;
            }
        ''')

        # Combo popups are separate Qt views. The application theme can style
        # them after the dialog stylesheet is applied, so pin their palette on
        # the popup itself to prevent black-on-black/white-on-white selections.
        for combo in (self._model_selector, self._date_selector):
            popup = combo.view()
            palette = popup.palette()
            palette.setColor(QPalette.ColorRole.Base, QColor('#ffffff'))
            palette.setColor(QPalette.ColorRole.Window, QColor('#ffffff'))
            palette.setColor(QPalette.ColorRole.Text, QColor('#202020'))
            palette.setColor(QPalette.ColorRole.WindowText, QColor('#202020'))
            palette.setColor(QPalette.ColorRole.Highlight, QColor('#2f79bd'))
            palette.setColor(QPalette.ColorRole.HighlightedText, QColor('#ffffff'))
            popup.setPalette(palette)
            popup.setItemDelegate(_ComboPopupDelegate(popup))
            popup.setStyleSheet('''
                QListView { background: #ffffff; color: #202020;
                    border: 1px solid #90989f; outline: 0; }
                QListView::item { background: #ffffff; color: #202020;
                    min-height: 24px; padding: 1px 5px; }
                QListView::item:hover { background: #dcebf8; color: #202020; }
                QListView::item:selected { background: #2f79bd; color: #ffffff; }
            ''')

    @staticmethod
    def _display_predictions(
            df: pd.DataFrame, include_market_details: bool = False,
    ) -> pd.DataFrame:
        # Keep the user-facing ticket view compact. Detailed model/market
        # probabilities remain in the internal report but are not repeated here.
        display = df.copy()

        def score_with_probability(
                score: str,
                probability: str,
                labels: dict[str, str] | None = None,
        ) -> pd.Series:
            values = display.get(score, pd.Series('', index=display.index)).fillna('').astype(str)
            probabilities = display.get(
                probability, pd.Series(float('nan'), index=display.index),
            )
            return pd.Series([
                (labels or {}).get(value, value) if pd.isna(prob)
                else f'{(labels or {}).get(value, value)}（{float(prob):.1%}）'
                for value, prob in zip(values, probabilities)
            ], index=display.index)

        def upset_direction(row: pd.Series) -> str:
            score = str(row.get('比分爆冷') or row.get('爆冷比分') or '')
            match = re.fullmatch(r'(\d+)\+?-(\d+)\+?', score)
            if match is None:
                return ''
            home, away = map(int, match.groups())
            if home > away:
                label, column = '主胜冷门', '模型主胜概率'
            elif home == away:
                label, column = '平局冷门', '模型平局概率'
            else:
                label, column = '客胜冷门', '模型客胜概率'
            probability = row.get(column)
            return label if pd.isna(probability) else f'{label}（{float(probability):.1%}）'

        def odds_pair(row: pd.Series, opening: str, current: str) -> str:
            def fmt(value):
                try:
                    return f'{float(value):.2f}' if pd.notna(value) else ''
                except (TypeError, ValueError):
                    return ''
            first, latest = fmt(row.get(opening)), fmt(row.get(current))
            if not first and not latest:
                return ''
            if not first:
                return f'现 {latest}'
            if not latest:
                return f'初 {first}'
            return f'初 {first} → 现 {latest}'

        def had_opening_current(row: pd.Series) -> str:
            pairs = [
                odds_pair(row, '首次采集胜奖金', '官方胜奖金'),
                odds_pair(row, '首次采集平奖金', '官方平奖金'),
                odds_pair(row, '首次采集负奖金', '官方负奖金'),
            ]
            return '｜'.join(pairs) if any(pairs) else '暂无指数'

        def hhad_opening_current(row: pd.Series) -> str:
            line_first = row.get('首次采集让球数')
            line_latest = row.get('官方让球数')
            def line(value):
                try:
                    return f'{float(value):g}' if pd.notna(value) else ''
                except (TypeError, ValueError):
                    return ''
            first, latest = line(line_first), line(line_latest)
            values = [
                odds_pair(row, '首次采集让胜奖金', '官方让胜奖金'),
                odds_pair(row, '首次采集让平奖金', '官方让平奖金'),
                odds_pair(row, '首次采集让负奖金', '官方让负奖金'),
            ]
            if not any(values) and not first and not latest:
                return '暂无指数'
            handicap_line = f'线 初 {first} → 现 {latest}' if first and latest else ''
            return '｜'.join(part for part in (handicap_line, *values) if part)

        def market_reference(row: pd.Series) -> str:
            accuracy = row.get('同阈值历史命中率')
            samples = row.get('筛选回测样本数')
            coverage = row.get('同阈值历史覆盖率')
            period = str(row.get('筛选回测期间') or '').strip()
            if pd.isna(accuracy):
                return '未提供可比回测'
            result = f'历史同档 {float(accuracy):.1%}'
            if pd.notna(coverage):
                result += f'｜覆盖 {float(coverage):.1%}'
            if pd.notna(samples):
                result += f'｜{int(samples)}场'
            if period and period.lower() != 'nan':
                result += f'｜{period}'
            return result

        def evidence_status(row: pd.Series) -> str:
            samples = pd.to_numeric(row.get('筛选回测样本数'), errors='coerce')
            concerns = []
            if pd.isna(samples):
                concerns.append('回测未披露')
            elif samples < 30:
                concerns.append('小样本')
            lineup = str(row.get('首发状态') or '')
            if lineup and lineup != '已确认':
                concerns.append('首发未确认')
            if str(row.get('模拟差异') or '').strip():
                concerns.append('模型分歧')
            gate = str(row.get('盘口门控') or '')
            if any(word in gate for word in ('冲突', '震荡', '不稳定')):
                concerns.append('市场信号不稳')
            return '需谨慎：' + '、'.join(concerns) if concerns else '证据项已披露'

        display['胜负首选'] = score_with_probability('胜平负首选', '胜平负首选概率')
        handicap_labels = {'胜': '让胜', '平': '让平', '负': '让负'}
        display['让球首选/次选'] = (
            score_with_probability('让球首选', '让球首选概率', handicap_labels) + '/'
            + score_with_probability('让球次选', '让球次选概率', handicap_labels)
        )
        display['总进球首选'] = score_with_probability(
            '竞彩总进球首选', '竞彩总进球首选概率',
        )
        display['半全场首选/次选'] = (
            score_with_probability('半全场首选', '半全场首选概率') + '/'
            + score_with_probability('半全场次选', '半全场次选概率')
        )
        def combined_scores(row: pd.Series) -> str:
            values = []
            for column in (
                '首选比分', '次选比分', '第三比分',
                '比分爆冷', '大小球进取比分',
            ):
                value = row.get(column)
                if pd.notna(value) and str(value).strip() and str(value).strip() not in values:
                    values.append(str(value).strip())
            return ' / '.join(values)

        display['比分情景（Top3/反向/高进球）'] = display.apply(
            combined_scores, axis=1,
        )
        display['置信度'] = display.get(
            '置信等级', pd.Series('', index=display.index),
        ).fillna('').astype(str)
        display['胜平负指数（首次采集→当前）'] = display.apply(had_opening_current, axis=1)
        display['让球指数（首次采集→当前）'] = display.apply(hhad_opening_current, axis=1)
        display['市场概率档参考'] = display.apply(market_reference, axis=1)
        display['分析依据'] = display.get(
            '预测依据', pd.Series('', index=display.index),
        ).fillna('').astype(str)
        display['胜负模型'] = display.get(
            '胜负模型类别', display.get(
                '模型类别', pd.Series('', index=display.index),
            ),
        ).fillna('').astype(str)
        display['模拟半全场'] = display.get(
            '模拟半全场', pd.Series('', index=display.index),
        ).fillna('').astype(str).map(
            lambda value: ' / '.join(
                part.strip() for part in value.split('/')[:2] if part.strip()
            ),
        )

        def simulation_difference(row: pd.Series) -> str:
            def first_choice(value: object) -> str:
                text = '' if pd.isna(value) else str(value).strip()
                return re.split(r'[\s/｜]', text, maxsplit=1)[0]

            def comparison(label: str, primary: str, simulated: str) -> str:
                if not primary or not simulated or primary == simulated:
                    return ''
                return f'{label} {primary}→{simulated}✕'

            parts = []
            primary_result = first_choice(row.get('胜平负首选'))[:1]
            simulated_result = first_choice(row.get('模拟胜负'))[:1]
            parts.append(comparison('胜负', primary_result, simulated_result))

            primary_handicap = first_choice(row.get('让球首选'))
            if primary_handicap in ('胜', '平', '负'):
                primary_handicap = f'让{primary_handicap}'
            simulated_handicap = first_choice(row.get('模拟让球'))
            parts.append(comparison('让球', primary_handicap, simulated_handicap))

            primary_total = first_choice(row.get('竞彩总进球首选'))
            simulated_total = first_choice(row.get('模拟竞彩总进球'))
            parts.append(comparison('总进球', primary_total, simulated_total))

            primary_score = str(row.get('首选比分') or '').strip()
            simulated_match = re.match(
                r'(\d+\+?-\d+\+?)', str(row.get('模拟Top3比分') or '').strip(),
            )
            simulated_score = simulated_match.group(1) if simulated_match else ''
            parts.append(comparison('比分', primary_score, simulated_score))

            primary_half_full = first_choice(row.get('半全场首选'))
            simulated_half_full = first_choice(row.get('模拟半全场'))
            parts.append(comparison('半全场', primary_half_full, simulated_half_full))
            return '｜'.join(part for part in parts if part)

        display['模拟差异'] = display.apply(simulation_difference, axis=1)
        display['证据状态'] = display.apply(evidence_status, axis=1)
        odds_series = read_odds_series()
        match_ids = display.get('比赛ID', pd.Series('', index=display.index))
        calculated_flow = [format_market_flow(mid, series=odds_series) for mid in match_ids]
        stored_gate = display.get('盘口门控', pd.Series('', index=display.index)).fillna('')
        display['盘口流向'] = [
            flow if flow not in ('', '待积累') else (gate or flow)
            for gate, flow in zip(stored_gate, calculated_flow)
        ]
        remaining = pd.to_numeric(
            display.get('距参考截止分钟', pd.Series(float('nan'), index=display.index)),
            errors='coerce',
        )
        display['距参考截止'] = remaining.map(
            lambda value: '' if pd.isna(value) else (
                f'约{float(value):.0f}分钟' if value >= 0 else '已过参考线'
            )
        )
        display['对阵'] = (
            display.get('主队', pd.Series('', index=display.index)).fillna('').astype(str)
            + ' vs '
            + display.get('客队', pd.Series('', index=display.index)).fillna('').astype(str)
        )

        def compact_direction(row: pd.Series) -> str:
            parts = []
            result = str(row.get('胜负首选') or '').strip()
            handicap = str(row.get('让球首选/次选') or '').strip()
            total = str(row.get('总进球首选') or '').strip()
            if result:
                parts.append(f'胜负 {result}')
            if handicap and handicap != '/':
                parts.append(f'让球 {handicap.split("/", 1)[0]}')
            if total:
                parts.append(f'进球 {total}')
            return '｜'.join(parts) or '数据待补'

        def compact_risk(row: pd.Series) -> str:
            parts = []
            disagreement = str(row.get('模拟差异') or '').strip()
            evidence = str(row.get('证据状态') or '').strip()
            lineup = str(row.get('阵容分析') or '').strip()
            if disagreement:
                parts.append(disagreement)
            if evidence.startswith('需谨慎'):
                parts.append(evidence.removeprefix('需谨慎：'))
            if lineup and lineup not in ('首发已确认', '阵容暂无明显影响'):
                parts.append(lineup)
            return '｜'.join(dict.fromkeys(parts)) or '正常'

        display['综合方向'] = display.apply(compact_direction, axis=1)
        display['风险提示'] = display.apply(compact_risk, axis=1)
        display['让球'] = display['让球首选/次选']
        display['总进球'] = display['总进球首选']
        display['半全场'] = display['半全场首选/次选']
        display['比分'] = display['比分情景（Top3/反向/高进球）']
        preferred = [
            '赛事编号', '比赛时间', '联赛', '对阵', '距参考截止',
            '综合方向', '盘口流向', '让球', '总进球', '半全场', '比分',
            '风险提示',
        ]
        if include_market_details:
            # The on-screen ticket stays compact, while the exported workbook
            # retains the recorded opening baseline and its movement to now.
            preferred[7:7] = [
                '胜平负指数（初盘/首次采集→当前）',
                '让球指数（初盘/首次采集→当前）',
            ]
            display['胜平负指数（初盘/首次采集→当前）'] = display[
                '胜平负指数（首次采集→当前）'
            ]
            display['让球指数（初盘/首次采集→当前）'] = display[
                '让球指数（首次采集→当前）'
            ]
        shown = display[[column for column in preferred if column in display.columns]].copy()
        for column in shown.columns:
            if (
                ('概率' in column or '优势' in column)
                and pd.api.types.is_numeric_dtype(shown[column])
            ):
                shown[column] = shown[column].map(
                    lambda value: '' if pd.isna(value) else f'{float(value):.1%}',
                )
        return shown

    def _load_saved_reports(self):
        def read_report(path: Path, columns: list[str]) -> pd.DataFrame:
            if not path.exists() or path.stat().st_size == 0:
                return pd.DataFrame(columns=columns)
            try:
                return pd.read_csv(path)
            except pd.errors.EmptyDataError:
                return pd.DataFrame(columns=columns)

        self._predictions = _sort_by_match_number(backfill_missing_simulations(
            read_report(PREDICTION_PATH, ['赛事编号', '联赛', '主队', '客队']),
        ))
        self._refresh_model_selector()
        self._refresh_date_selector()
        self._refresh_advice_selector()
        self._render()

    def _refresh_model_selector(self):
        had_selection = self._model_selector.count() > 0
        current = self._model_selector.currentData() if had_selection else None
        active = _upcoming_predictions(self._predictions)
        scopes = _prediction_model_scopes(active)
        reported_leagues = [scope for scope in scopes.unique() if scope]
        leagues = list(dict.fromkeys(self._trained_leagues + sorted(reported_leagues)))
        dedicated_count = int(scopes.ne('').sum())
        generic_count = int(scopes.eq('').sum())
        counts = scopes.value_counts().to_dict()
        self._model_selector.blockSignals(True)
        self._model_selector.clear()
        self._model_selector.addItem(
            f'全部模型结果（{len(active)}场，含不同模型）', ALL_MODELS,
        )
        self._model_selector.addItem(
            f'全部专用模型（{dedicated_count}场）', DEDICATED_MODELS,
        )
        self._model_selector.addItem(
            f'通用/市场模型（{generic_count}场）', GENERIC_MODELS,
        )
        simulated = pd.to_numeric(
            active.get(
                '模拟次数', pd.Series(0, index=active.index),
            ), errors='coerce',
        ).fillna(0).gt(0)
        self._model_selector.addItem(
            f'蒙特卡洛模拟（{int(simulated.sum())}场）', SIMULATION_MODELS,
        )
        for league in leagues:
            self._model_selector.addItem(
                f'{league}专用模型（{int(counts.get(league, 0))}场）', league,
            )
        if current is None:
            # Opening the window must show the complete report.  Defaulting to
            # a model subset made valid rows look as if they had not loaded.
            current = ALL_MODELS
        index = self._model_selector.findData(current)
        self._model_selector.setCurrentIndex(index if index >= 0 else 0)
        self._model_selector.blockSignals(False)

    def _visible_predictions(self) -> pd.DataFrame:
        visible = filter_predictions_by_model(
            self._predictions,
            self._model_selector.currentData() or ALL_MODELS,
        )
        visible = _upcoming_predictions(visible)
        selected_date = self._date_selector.currentData()
        if selected_date and '比赛时间' in visible.columns:
            dates = visible['比赛时间'].fillna('').astype(str).str.slice(0, 10)
            visible = visible.loc[dates.eq(str(selected_date))]
        return _sort_by_match_number(visible)

    def _refresh_advice_selector(self):
        current = self._advice_selector.currentData()
        active = _upcoming_predictions(self._predictions)
        advice = active.get(
            '建议状态', pd.Series('', index=active.index),
        ).fillna('').astype(str)
        selected = int(advice.eq('精选主推').sum())
        high = int(advice.eq('高置信主推').sum())
        self._advice_selector.blockSignals(True)
        self._advice_selector.clear()
        self._advice_selector.addItem(f'达到筛选阈值（{selected + high}场）', '正式推荐')
        self._advice_selector.addItem(f'A档概率样本（{selected}场）', '精选主推')
        self._advice_selector.addItem(f'B档概率样本（{high}场）', '高置信主推')
        self._advice_selector.addItem(f'全部场次（{len(active)}场）', '')
        if current is None or (current == '正式推荐' and selected + high == 0):
            # Never open on an empty recommendation view when the report itself
            # contains fixtures.  Users can still select each recommendation
            # grade explicitly when such rows exist.
            current = ''
        index = self._advice_selector.findData(current)
        self._advice_selector.setCurrentIndex(index if index >= 0 else 0)
        self._advice_selector.blockSignals(False)

    def _refresh_date_selector(self):
        current = self._date_selector.currentData()
        upcoming = _upcoming_predictions(self._predictions)
        if '比赛时间' in upcoming.columns:
            dates = upcoming['比赛时间'].fillna('').astype(str).str.slice(0, 10)
            dates = sorted(value for value in dates.unique() if re.fullmatch(r'\d{4}-\d{2}-\d{2}', value))
        else:
            dates = []
        self._date_selector.blockSignals(True)
        self._date_selector.clear()
        upcoming_count = len(_upcoming_predictions(self._predictions))
        self._date_selector.addItem(f'尚未开赛（{upcoming_count}场）', '')
        for value in dates:
            count = int(upcoming['比赛时间'].astype(str).str.startswith(value).sum())
            weekday = '一二三四五六日'[date.fromisoformat(value).weekday()]
            self._date_selector.addItem(f'{value} 周{weekday}（{count}场）', value)
        index = self._date_selector.findData(current)
        self._date_selector.setCurrentIndex(index if index >= 0 else 0)
        self._date_selector.blockSignals(False)

    def _render(self):
        visible = self._visible_predictions()
        active_count = len(_upcoming_predictions(self._predictions))
        display_frame = self._display_predictions(visible)
        priorities = _daily_priority_aspects(visible).reset_index(drop=True)
        marked_cells = _mark_priority_cells(display_frame, priorities)
        if self._table is None:
            self._table = ExcelTable(
                parent=self,
                df=display_frame,
                readonly=True,
                supports_sorting=True,
            )
            self._table_layout.addWidget(self._table)
        else:
            self._table.update_dataframe(display_frame)
        table = self._table
        for row_index, columns in marked_cells.items():
            for column in columns:
                column_index = display_frame.columns.get_loc(column)
                item = table.item(row_index, column_index)
                if item is None:
                    continue
                item.setForeground(QBrush(QColor('#c62828')))
                item.setBackground(QBrush(QColor('#fff1f1')))
                font = QFont(item.font())
                font.setBold(True)
                item.setFont(font)
        difference_column = '模拟差异'
        if difference_column in display_frame.columns:
            column_index = display_frame.columns.get_loc(difference_column)
            for row_index, value in enumerate(display_frame[difference_column].fillna('')):
                item = table.item(row_index, column_index)
                if item is None:
                    continue
                text = str(value)
                if '✕' in text:
                    item.setForeground(QBrush(QColor('#c62828')))
                    item.setBackground(QBrush(QColor('#fff1f1')))
                elif '✓' in text:
                    item.setForeground(QBrush(QColor('#137333')))
                    item.setBackground(QBrush(QColor('#eef8f0')))
                font = QFont(item.font())
                font.setBold(True)
                item.setFont(font)
        table.horizontalHeader().setSectionsMovable(True)
        table.horizontalHeader().setFixedHeight(26)
        table.verticalHeader().setDefaultSectionSize(25)
        table.verticalHeader().setMinimumSectionSize(23)
        priority_text = _priority_summary(_upcoming_predictions(self._predictions))
        if visible.empty and not self._predictions.empty:
            self._summary.setText(
                f'当前显示 0 场  ·  未开赛 {active_count} 场  ·  {priority_text}'
            )
        else:
            self._summary.setText(
                f'当前显示 {len(visible)} 场  ·  未开赛 {active_count} 场  ·  {priority_text}'
            )

    def _sync(self):
        if self._lineup_running:
            QMessageBox.information(self, '请稍候', '阵容督导正在更新，完成后即可手动同步。')
            return
        self._foreground_sync_running = True
        self._lineup_timer.stop()
        runner = TaskRunnerDialog(
            title='同步竞猜数据',
            info='正在获取官方场次并生成预测…',
            task_fn=run_daily_sporttery,
            parent=self,
        )
        try:
            result = runner.run()
        finally:
            self._foreground_sync_running = False
            self._schedule_lineup_supervision()
        if runner.error_message is not None or result is None:
            QMessageBox.critical(
                self, '同步失败', runner.error_message or '同步任务没有返回数据，请稍后重试。',
            )
            return
        self._predictions, _ = result
        _save_daily_recommendation_snapshot(
            build_daily_recommendations(self._predictions),
        )
        self._refresh_model_selector()
        self._refresh_date_selector()
        self._refresh_advice_selector()
        self._render()
        learning_message = ''
        if LEARNING_STATUS_PATH.exists():
            try:
                learning = json.loads(LEARNING_STATUS_PATH.read_text(encoding='utf-8'))
                learning_message = (
                    f'\n每日复盘：本次结算 {int(learning.get("newly_settled") or 0)} 场，'
                    f'累计 {int(learning.get("settled_samples") or 0)} 场，'
                    f'等待官方赛果 {int(learning.get("pending_samples") or 0)} 场；'
                    f'有效训练样本 {int(learning.get("model_samples") or 0)} 场；'
                    f'{learning.get("model_status", "积累样本中")}。'
                )
                next_training = int(learning.get('next_training_at') or 0)
                model_samples = int(learning.get('model_samples') or 0)
                if next_training > model_samples:
                    learning_message += f' 距下轮自动训练还差{next_training-model_samples}场。'
            except (OSError, ValueError, TypeError):
                pass
        QMessageBox.information(
            self, '同步完成',
            f'已生成 {len(_upcoming_predictions(self._predictions))} 场未开赛预测。'
            f'{learning_message}',
        )

    def _review_history(self):
        runner = TaskRunnerDialog(
            title='补同步历史赛果',
            info='正在扫描遗漏场次、获取官方赛果并执行复盘…',
            task_fn=lambda: review_and_learn(full_backfill=True),
            parent=self,
        )
        result = runner.run()
        if runner.error_message is not None or result is None:
            QMessageBox.critical(
                self, '补同步失败',
                runner.error_message or '复盘任务没有返回结果，请稍后重试。',
            )
            return
        self._render()
        accuracy = result.get('result_accuracy')
        accuracy_text = f'{float(accuracy):.1%}' if accuracy is not None else '--'
        newly_settled = int(result.get('newly_settled') or 0)
        message = (
            ('本次未发现新增赛果；以下均为累计复盘数据。\n' if newly_settled == 0 else '')
            + f'本次补结算 {newly_settled} 场，'
            f'累计复盘 {int(result.get("settled_samples") or 0)} 场，'
            f'等待官方赛果 {int(result.get("pending_samples") or 0)} 场，'
            f'胜平负历史命中率 {accuracy_text}（累计已结算样本）。\n'
            f'本次补入官方市场样本 {int(result.get("new_official_history") or 0)} 场，'
            f'累计收集 {int(result.get("total_training_samples") or 0)} 场，'
            f'其中有效训练样本 {int(result.get("model_samples") or 0)} 场。\n'
            f'模型状态：{result.get("model_status", "积累样本中")}。\n'
            f'模型迭代：已审计 {int(result.get("evolution_attempts") or 0)} 次，'
            f'当前采用第 {int(result.get("champion_generation") or 0)} 代。'
        )
        next_training = int(result.get('next_training_at') or 0)
        model_samples = int(result.get('model_samples') or 0)
        if next_training > model_samples:
            message += (
                f'\n自动训练：还需新增 {next_training - model_samples} 场有效完场样本，'
                '达到后自动启动冠军/挑战者时间外审计。'
            )
        selection_rows = (result.get('selection_profile') or {}).get('rows') or []
        recommended = [
            row for row in selection_rows
            if row.get('grade') in ('精选主推', '高置信主推')
        ]
        if recommended:
            message += '\n历史筛选阈值：' + '；'.join(
                f'{"A档" if row["grade"] == "精选主推" else "B档"}'
                f'≥{float(row["threshold"]):.1%}'
                f'（历史同档{float(row["accuracy"]):.1%}）'
                for row in recommended
            )
        over_under_rows = (result.get('over_under_profile') or {}).get('directions') or []
        if over_under_rows:
            message += '\n大小球滚动门控：' + '；'.join(
                f'{row.get("label", row.get("pick", ""))}'
                f'{"启用" if row.get("enabled") else "暂停"}'
                f'（门槛≥{float(row.get("threshold") or 0):.0%}，'
                f'留出{int(row.get("audit_samples") or 0)}场/'
                f'{float(row.get("audit_accuracy") or 0):.1%}）'
                for row in over_under_rows
            )
        market_accuracy = result.get('accuracy_by_market') or {}
        available_accuracy = [
            f'{name}{float(value):.1%}'
            for name, value in market_accuracy.items() if value is not None
        ]
        if available_accuracy:
            message += '\n各玩法冻结预测复盘：' + '；'.join(available_accuracy)
        model_lines = []
        for name, audit in result.get('accuracy_by_model', {}).items():
            samples = int(audit.get('samples') or 0)
            if samples < 10:
                model_lines.append(f'{name}：{samples}场，样本不足，继续积累')
            else:
                model_lines.append(
                    f'{name}：{samples}场，'
                    f'模型 {float(audit.get("accuracy") or 0):.1%}，'
                    f'市场基线 {float(audit.get("market_accuracy") or 0):.1%}，'
                    f'{audit.get("status", "继续观察")}'
                )
        if model_lines:
            message += '\n\n分模型最近实战：\n' + '\n'.join(model_lines)
        if result.get('review_error'):
            QMessageBox.warning(
                self, '部分赛果暂未补齐',
                f'{message}\n\n'
                f'官方赛果服务暂时无响应，已保留现有复盘数据。'
                f'等待补齐的 {int(result.get("pending_samples") or 0)} 场将在下次自动重试；'
                f'不影响当前预测和已训练模型。',
            )
        else:
            QMessageBox.information(self, '补同步完成', message)

    def _show_yesterday_hits(self):
        # Reading the local settlement cache is intentionally instant and never
        # starts a network request. Use the existing review button when official
        # results still need to be synchronized.
        if self._yesterday_dialog is not None and self._yesterday_dialog.isVisible():
            self._yesterday_dialog.raise_()
            self._yesterday_dialog.activateWindow()
            return
        self._yesterday_dialog = YesterdayHitDetailsDialog(self)
        self._yesterday_dialog.show()
        self._yesterday_dialog.raise_()
        self._yesterday_dialog.activateWindow()

    def _show_market_trends(self):
        visible = self._visible_predictions()
        if visible.empty and self._predictions.empty:
            QMessageBox.information(self, '没有数据', '请先同步竞猜数据。')
            return
        source = visible if not visible.empty else self._predictions
        if self._market_dialog is not None and self._market_dialog.isVisible():
            self._market_dialog.raise_()
            self._market_dialog.activateWindow()
            return
        self._market_dialog = MarketTrendDialog(source, self)
        self._market_dialog.show()
        self._market_dialog.raise_()
        self._market_dialog.activateWindow()

    def _show_daily_recommendations(self):
        if self._predictions.empty:
            QMessageBox.information(self, '没有数据', '请先同步竞猜数据。')
            return
        if (
                self._daily_recommendations_dialog is not None
                and self._daily_recommendations_dialog.isVisible()
        ):
            self._daily_recommendations_dialog.raise_()
            self._daily_recommendations_dialog.activateWindow()
            return
        self._daily_recommendations_dialog = DailyRecommendationsDialog(
            self._predictions, self,
        )
        self._daily_recommendations_dialog.show()
        self._daily_recommendations_dialog.raise_()
        self._daily_recommendations_dialog.activateWindow()

    def _export(self):
        visible = self._visible_predictions()
        if visible.empty:
            QMessageBox.information(self, '没有数据', '请先同步竞猜数据。')
            return
        selected = self._model_selector.currentText().split('（', 1)[0]
        exported_at = datetime.now()
        timestamp = exported_at.strftime('%Y-%m-%d_%H-%M-%S-%f')[:-3]
        path = (
            prediction_export_root()
            / f'{timestamp}-{_safe_filename(selected)}.xlsx'
        )
        try:
            exported = self._display_predictions(
                visible, include_market_details=True,
            ).reset_index(drop=True)
            priorities = _daily_priority_aspects(visible).reset_index(drop=True)
            _mark_priority_cells(exported, priorities)
            write_predictions_xlsx(exported, path)
        except PermissionError:
            QMessageBox.critical(
                self, '导出失败', 'Excel 文件正在被占用，请关闭对应文件后重试。',
            )
            return
        except Exception as error:
            QMessageBox.critical(self, '导出失败', f'无法生成 Excel：{error}')
            return
        QMessageBox.information(
            self, '导出完成', f'已生成 Excel 文件：\n{display_export_path(path)}',
        )
