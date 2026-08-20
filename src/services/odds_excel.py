"""Export recorded odds history and predictions as readable Excel workbooks."""

import os
from datetime import datetime, timedelta, timezone
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from src.services.odds_tracking import (
    format_match_drift, intent_for_match, read_odds_series,
)

FILE_NAME = 'ProphitBet-赔率记录.xlsx'


def default_output_path(file_name: str = FILE_NAME) -> Path:
    """Prefer the visible Windows desktop, falling back to the WSL home."""
    candidates = [
        Path('/mnt/c/Users/Administrator/Desktop'),
        Path.home() / 'Desktop',
        Path.home(),
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate / file_name
    return Path.cwd() / file_name


def _local_time(value: str) -> str:
    try:
        return datetime.fromisoformat(value).astimezone(
            timezone(timedelta(hours=8)),
        ).strftime('%Y-%m-%d %H:%M')
    except (TypeError, ValueError):
        return str(value or '')


def _display_width(value) -> int:
    text = str(value if value is not None else '')
    return sum(2 if ord(char) > 127 else 1 for char in text)


HEADERS = [
    '记录时间', '赛事编号', '联赛', '主队', '客队',
    '主胜赔率', '平局赔率', '客胜赔率', '赔率变动', '庄家意图',
    '让球数', '让胜', '让平', '让负', '官方更新时间',
]


def export_odds_history_excel(output: Path = None) -> tuple[int, Path]:
    """Write all recorded odds observations to Excel; return (rows, path)."""
    output = output or default_output_path()
    series = read_odds_series()
    rows = []
    for match_id, observations in series.items():
        for obs in observations:
            had = obs.get('had') or {}
            hhad = obs.get('hhad') or {}
            rows.append([
                _local_time(obs.get('captured_at', '')),
                obs.get('match_num', ''),
                obs.get('league', ''),
                obs.get('home', ''),
                obs.get('away', ''),
                had.get('H'), had.get('D'), had.get('A'),
                format_match_drift(match_id, series=series),
                intent_for_match(match_id, had, series=series),
                hhad.get('line'), hhad.get('H'), hhad.get('D'), hhad.get('A'),
                obs.get('had_update', ''),
            ])
    rows.sort(key=lambda row: (str(row[0]), str(row[1])))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = '赔率记录'
    sheet.append(HEADERS)
    for row in rows:
        sheet.append(row)
    _style_sheet(sheet, len(rows), money_columns=(6, 7, 8, 12, 13, 14))
    return len(rows), _save_workbook(workbook, output)


def _style_sheet(sheet, row_count: int, money_columns=()):
    header_fill = PatternFill('solid', start_color='4472C4')
    for column in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=1, column=column)
        cell.font = Font(bold=True, color='FFFFFF', name='宋体')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for r in range(2, row_count + 2):
        for c in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=r, column=c)
            cell.font = Font(name='宋体')
            cell.alignment = Alignment(vertical='center')
            if c in money_columns:
                cell.number_format = '0.00'

    for column in range(1, sheet.max_column + 1):
        longest = max(
            (
                _display_width(sheet.cell(row=r, column=column).value)
                for r in range(1, row_count + 2)
            ),
            default=8,
        )
        sheet.column_dimensions[
            sheet.cell(row=1, column=column).column_letter
        ].width = max(8, min(longest + 3, 40))
    sheet.row_dimensions[1].height = 30
    for r in range(2, row_count + 2):
        sheet.row_dimensions[r].height = 20
    sheet.freeze_panes = 'A2'


def _save_workbook(workbook: Workbook, output: Path) -> Path:
    try:
        workbook.save(output)
    except PermissionError:
        # The target file is probably open in Excel; use a timestamped name.
        timestamp = datetime.now(timezone(timedelta(hours=8))).strftime(
            '%Y%m%d-%H%M',
        )
        output = output.with_name(f'{output.stem}-{timestamp}{output.suffix}')
        workbook.save(output)
    return output


def export_dataframe_excel(frame: pd.DataFrame, file_name: str) -> tuple[int, Path]:
    """Write a DataFrame to a styled Excel file on the desktop."""
    output = default_output_path(file_name)
    rows = frame.fillna('').values.tolist()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = '预测'
    sheet.append([str(column) for column in frame.columns])
    for row in rows:
        sheet.append(row)
    money_columns = tuple(
        index + 1 for index, column in enumerate(frame.columns)
        if any(word in str(column) for word in ('赔率', '奖金', '概率', '优势'))
    )
    _style_sheet(sheet, len(rows), money_columns=money_columns)
    return len(rows), _save_workbook(workbook, output)
