import pandas as pd
from PySide6.QtWidgets import QApplication, QDialog, QPushButton

from src.gui.i18n import translate_widget
from src.gui import main as main_window
from src.gui.widgets.tables import DataFrameTable, DataFrameTableModel, ExcelTable
from src.gui.windows.models import fixtures
from src.gui.windows.models.predictor import PredictorDialog
from src.gui.windows import sporttery as sporttery_window
from src.gui.windows.sporttery import (
    DEDICATED_MODELS, GENERIC_MODELS, SIMULATION_MODELS, filter_predictions_by_model,
    write_predictions_xlsx,
)
from src.network.fixtures.utils import match_fixture_teams
from src.preprocessing.utils.inputs import construct_inputs_by_teams


_APP = None


def test_lineup_supervisor_wakes_before_matches_enter_window(monkeypatch):
    from types import SimpleNamespace
    calls = []
    timer = SimpleNamespace(stop=lambda: None, start=calls.append)
    window = SimpleNamespace(
        _lineup_timer=timer,
        _predictions=pd.DataFrame([{'比赛时间': '2099-09-05 18:00', '首发状态': '未获取'}]),
    )
    monkeypatch.setattr(sporttery_window, 'lineup_api_configured', lambda: True)
    sporttery_window.SportteryPredictionsDialog._schedule_lineup_supervision(window)
    assert calls and 0 < calls[-1] <= 15 * 60 * 1000


def _app():
    global _APP
    _APP = QApplication.instance() or QApplication([])
    return _APP


def test_daily_recommendations_keeps_half_time_review_inside_yesterday_review(monkeypatch):
    _app()
    monkeypatch.setattr(
        sporttery_window, 'build_daily_recommendations',
        lambda predictions: pd.DataFrame(),
    )
    monkeypatch.setattr(
        sporttery_window, '_save_daily_recommendation_snapshot', lambda frame: None,
    )
    monkeypatch.setattr(
        sporttery_window, 'build_half_time_observations',
        lambda predictions: pd.DataFrame(),
    )
    monkeypatch.setattr(
        sporttery_window, '_save_half_time_observation_snapshot', lambda frame: None,
    )

    dialog = sporttery_window.DailyRecommendationsDialog(pd.DataFrame())
    labels = {button.text() for button in dialog.findChildren(QPushButton)}

    assert '昨日推荐复盘' in labels
    assert '半场推荐复核' not in labels
    assert '半场组合账本' in labels


def test_core_prediction_window_is_reused_without_resurfacing_main(monkeypatch):
    class Signal:
        def connect(self, callback):
            self.callback = callback

    class FakeDialog:
        def __init__(self):
            self.finished = Signal()
            self.visible = False
            self.raised = self.activated = 0

        def setAttribute(self, *args):
            pass

        def isVisible(self):
            return self.visible

        def show(self):
            self.visible = True

        def raise_(self):
            self.raised += 1

        def activateWindow(self):
            self.activated += 1

    created = []
    monkeypatch.setattr(
        main_window, 'SportteryPredictionsDialog',
        lambda: created.append(FakeDialog()) or created[-1],
    )

    class Owner:
        _sporttery_dialog = None
        hidden = 0

        def hide(self):
            self.hidden += 1

        def _restore_main_after_sporttery(self):
            pass

    owner = Owner()
    main_window.MainWindow._open_sporttery_predictions(owner)
    main_window.MainWindow._open_sporttery_predictions(owner)

    assert len(created) == 1
    assert owner.hidden == 1
    assert created[0].raised == 2


def test_large_readonly_table_uses_lazy_dataframe_model():
    _app()
    frame = pd.DataFrame({
        'Home': [f'Home {index}' for index in range(8000)],
        'Away': [f'Away {index}' for index in range(8000)],
        '1': [2.0] * 8000,
    })
    table = DataFrameTable(parent=None, df=frame)

    assert isinstance(table.model(), DataFrameTableModel)
    assert table.model().rowCount() == 8000
    hits = table.compute_matches('Home 7999', exact_match=True, selected_scope_index=1)
    assert len(hits) == 1
    assert hits[0].column() == 0
    assert hits[0].data() == 'Home 7999'


def test_ticket_header_sort_uses_card_date_and_natural_sequence():
    _app()
    frame = pd.DataFrame([
        {'赛事编号': '周六001', '比赛时间': '2026-08-29 17:00'},
        {'赛事编号': '周五010', '比赛时间': '2026-08-29 03:00'},
        {'赛事编号': '周五002', '比赛时间': '2026-08-29 00:30'},
    ])
    model = DataFrameTableModel(frame)
    model.sort(0)
    assert model._df['赛事编号'].tolist() == ['周五002', '周五010', '周六001']

    widget = ExcelTable(None, frame, supports_query_search=False)
    widget.sortItems(0)
    assert [widget.item(row, 0).text() for row in range(widget.rowCount())] == [
        '周五002', '周五010', '周六001',
    ]


def test_excel_table_reuses_widget_when_dataframe_changes():
    _app()
    widget = ExcelTable(
        None, pd.DataFrame([{'赛事编号': '周五001', '结果': '胜'}]),
        supports_query_search=False,
    )
    identity = id(widget)

    widget.update_dataframe(pd.DataFrame([
        {'赛事编号': '周六002', '结果': '平'},
        {'赛事编号': '周六003', '结果': '负'},
    ]))

    assert id(widget) == identity
    assert widget.rowCount() == 2
    assert widget.columnCount() == 2
    assert widget.item(0, 1).text() == '平'


def test_fixture_worker_maps_official_sporttery_rows(monkeypatch):
    class FakeClient:
        def __init__(self, *args, **kwargs):
            pass

        def selling_matches(self):
            return [{
                'matchDate': '2026-08-11',
                'leagueAllName': '瑞典超级联赛',
                'homeTeamAllName': '天狼星',
                'awayTeamAllName': '布鲁马波卡纳',
                'had': {'h': '1.20', 'd': '5.65', 'a': '8.40'},
            }]

    monkeypatch.setattr(fixtures, 'SportteryMobileClient', FakeClient)
    worker = fixtures.FixtureFetchWorker(
        fixture_url='https://invalid.example',
        date_str='Aug 11',
        iso_date='2026-08-11',
        league_id='瑞超',
    )

    frame = worker._sporttery_fixtures()
    assert frame.to_dict('records') == [{
        'Home': 'Sirius',
        'Away': 'Brommapojkarna',
        '1': 1.2,
        'X': 5.65,
        '2': 8.4,
    }]


def test_translation_keeps_native_window_title_ascii():
    _app()
    dialog = QDialog()
    dialog.setWindowTitle('Fixtures Dialog')
    translate_widget(dialog)
    assert dialog.windowTitle() == 'Fixtures Dialog'
    assert dialog.windowTitle().isascii()


def test_new_team_uses_neutral_league_statistics_without_renaming():
    history = pd.DataFrame([
        {
            'Date': '2026-05-24', 'Season': 2025, 'Week': 38,
            'Home': 'Arsenal', 'Away': 'Chelsea', '1': 1.8, 'X': 3.5, '2': 4.5,
            'HPoints': 2.4, 'APoints': 1.5,
        },
        {
            'Date': '2026-05-17', 'Season': 2025, 'Week': 37,
            'Home': 'Liverpool', 'Away': 'Arsenal', '1': 2.0, 'X': 3.4, '2': 3.8,
            'HPoints': 2.0, 'APoints': 1.7,
        },
    ])
    fixture = pd.DataFrame([{
        'Home': 'Coventry', 'Away': 'Arsenal', '1': 3.1, 'X': 3.2, '2': 2.4,
    }])

    built = construct_inputs_by_teams(history, fixture)

    assert built.at[0, 'Home'] == 'Coventry'
    assert built.at[0, 'Away'] == 'Arsenal'
    assert built.at[0, 'HPoints'] == 2.2
    assert not built.isna().any().any()


def test_unknown_fixture_team_is_not_fuzzily_replaced_by_another_club():
    fixtures_df = pd.DataFrame([{
        'Home': 'Coventry', 'Away': 'Arsenal', '1': 3.1, 'X': 3.2, '2': 2.4,
    }])
    history = pd.DataFrame({
        'Home': ['Everton', 'Arsenal'],
        'Away': ['Chelsea', 'Liverpool'],
    })

    matched = match_fixture_teams(fixtures_df, history)

    assert matched.at[0, 'Home'] == 'Coventry'
    assert matched.at[0, 'Away'] == 'Arsenal'


def test_manual_prediction_allows_every_historical_team_in_either_role():
    class FakeModelDatabase:
        league_id = '英超'

        @staticmethod
        def get_model_ids():
            return []

    _app()
    history = pd.DataFrame({
        'Home': ['Arsenal', 'Liverpool'],
        'Away': ['Chelsea', 'Everton'],
    })
    dialog = PredictorDialog(history, FakeModelDatabase())

    expected = {'Arsenal', 'Liverpool', 'Chelsea', 'Everton', 'Coventry'}
    assert set(dialog._home_teams) == expected
    assert set(dialog._away_teams) == expected
    dialog.close()


def test_manual_prediction_ranks_models_by_holdout_not_training_accuracy():
    class ModelContext:
        _model_quality_key = PredictorDialog._model_quality_key

    dialog = ModelContext()
    dialog._model_configs = {
        '训练看起来很高': {
            'train': {'results': {'fit': {'Accuracy': 0.99}}},
        },
        '独立测试更好': {
            'train': {'tuning': {'test_accuracy': 0.56, 'test_samples': 400}},
        },
        '独立测试较低': {
            'train': {'tuning': {'test_accuracy': 0.52, 'test_samples': 800}},
        },
    }
    ranked = sorted(dialog._model_configs, key=dialog._model_quality_key, reverse=True)

    assert ranked == ['独立测试更好', '独立测试较低', '训练看起来很高']


def test_fixture_model_picker_uses_holdout_ranking():
    class ModelContext:
        _model_quality_key = fixtures.FixturesDialog._model_quality_key

    dialog = ModelContext()
    dialog._model_configs = {
        '无独立测试': {'train': {'results': {'fit': {'Accuracy': 1.0}}}},
        '稳定模型': {'train': {'tuning': {'test_accuracy': 0.55, 'test_samples': 600}}},
        '较弱模型': {'train': {'tuning': {'test_accuracy': 0.51, 'test_samples': 900}}},
    }

    ranked = sorted(dialog._model_configs, key=dialog._model_quality_key, reverse=True)

    assert ranked == ['稳定模型', '较弱模型', '无独立测试']


def test_sporttery_dedicated_model_filter_never_mixes_leagues():
    predictions = pd.DataFrame([
        {'赛事编号': '周六001', '联赛': '英超', '专用模型联赛': '英超'},
        {'赛事编号': '周六002', '联赛': '西甲', '专用模型联赛': '西甲'},
        {'赛事编号': '周六003', '联赛': '欧冠', '专用模型联赛': ''},
    ])

    english = filter_predictions_by_model(predictions, '英超')
    dedicated = filter_predictions_by_model(predictions, DEDICATED_MODELS)
    generic = filter_predictions_by_model(predictions, GENERIC_MODELS)

    assert english['赛事编号'].tolist() == ['周六001']
    assert dedicated['赛事编号'].tolist() == ['周六001', '周六002']
    assert generic['赛事编号'].tolist() == ['周六003']


def test_sporttery_selector_lists_only_complete_dedicated_model_sets(monkeypatch):
    class FakeModelDatabase:
        def __init__(self, _league):
            self.index = {
                '英超': {
                    '英超胜平负模型': {}, '英超大小球模型': {},
                    '英超比分模型': {}, '英超半全场模型': {},
                },
                '西甲': {
                    '西甲胜平负模型': {}, '西甲比分模型': {},
                },
            }

    monkeypatch.setattr(sporttery_window, 'ModelDatabase', FakeModelDatabase)

    assert sporttery_window.trained_dedicated_leagues() == ['英超']


def test_sporttery_excel_export_is_a_real_formatted_workbook(tmp_path):
    from openpyxl import load_workbook

    path = tmp_path / '2026-08-15-英超专用模型.xlsx'
    write_predictions_xlsx(pd.DataFrame([{
        '赛事编号': '周六001', '联赛': '英超', '胜负首选': '胜（55.0%）',
    }]), path)

    workbook = load_workbook(path)
    sheet = workbook['竞彩预测']
    assert sheet.freeze_panes == 'A2'
    assert sheet.auto_filter.ref == 'A1:C2'
    assert sheet['A1'].value == '赛事编号'
    assert sheet['A2'].value == '周六001'


def test_sporttery_excel_export_highlights_only_priority_cells(tmp_path):
    from openpyxl import load_workbook

    path = tmp_path / 'priority.xlsx'
    write_predictions_xlsx(pd.DataFrame([{
        '赛事编号': '周六001', '综合方向': '★胜负重点｜胜负 胜（66.0%）',
    }]), path)

    sheet = load_workbook(path)['竞彩预测']
    assert sheet['A2'].fill.fill_type is None
    assert sheet['B2'].font.color.rgb == '00C62828'
    assert sheet['B2'].fill.fgColor.rgb == '00FFF1F1'
    assert sheet['B2'].font.bold is True


def test_priority_marker_is_added_only_to_the_selected_market_cells():
    display = pd.DataFrame([{
        '综合方向': '胜负 胜（66.0%）', '让球': '让胜（62.0%）',
        '大小球': '大于2.5球（61.0%）', '半全场': '胜胜（36.0%）',
        '比分': '2-1 / 1-0 / 2-0',
    }])
    marked = sporttery_window._mark_priority_cells(
        display, pd.Series([['让球', '比分']]),
    )
    assert marked == {0: ['让球', '比分']}
    assert display.loc[0, '让球'].startswith('★让球重点｜')
    assert display.loc[0, '比分'].startswith('★比分重点｜')
    assert not display.loc[0, '综合方向'].startswith('★')
    assert not display.loc[0, '大小球'].startswith('★')


def test_daily_recommendations_include_calibrated_draw_and_structural_handicap(monkeypatch):
    predictions = pd.DataFrame([{
        '赛事编号': '周六001', '比赛时间': '2099-08-29 18:00',
        '联赛': '英超', '主队': '阿森纳', '客队': '切尔西',
        '建议状态': '精选主推', '盘口门控': '稳定',
        '胜平负首选': '胜', '胜平负首选概率': 0.66,
        '模型主胜概率': 0.66, '模型平局概率': 0.22, '模型客胜概率': 0.12,
        '首次采集胜奖金': 1.90, '首次采集平奖金': 3.40, '首次采集负奖金': 4.20,
        '官方胜奖金': 1.80, '官方平奖金': 3.50, '官方负奖金': 4.40,
        '模拟胜负': '胜 68.0%',
        '官方让球数': -1, '让球首选': '胜', '让球首选概率': 0.64,
        '模型让胜概率': 0.64, '模型让平概率': 0.20, '模型让负概率': 0.16,
        '首次采集让球数': -1,
        '首次采集让胜奖金': 2.10, '首次采集让平奖金': 3.10,
        '首次采集让负奖金': 3.20,
        '官方让胜奖金': 2.0, '官方让平奖金': 3.2, '官方让负奖金': 3.4,
        '模拟让球': '让胜 65.0%',
        '让球最大概率优势': 0.05,
        '大小球首选': '大于2.5球', '大小球首选概率': 0.61,
        '模拟竞彩总进球': '3球', '模拟竞彩总进球概率': 0.27,
        '半全场首选': '胜胜', '半全场首选概率': 0.36,
        '半全场次选': '平胜', '半全场次选概率': 0.22,
        '比分推荐状态': '推荐', '首选比分': '2-1',
        '首选比分概率': 0.16, '次选比分': '1-0', '次选比分概率': 0.13,
        '第三比分': '2-0',
        '原始最高概率比分概率': 0.16,
    }])
    result = sporttery_window.build_daily_recommendations(predictions)
    assert len(result) == 1
    assert result.loc[0, '推荐玩法'] == '让球胜平负'
    assert result.loc[0, '重点选项'] == '★ -1球 胜'
    assert result.loc[0, '正式模型概率'] == '64.0%'
    assert result.loc[0, '推荐等级'] == '核心重点'
    assert result.loc[0, '价值评估'] == 'SP 2.00｜保守概率 60.0%｜EV +20.0%'
    assert result.loc[0, '比分参考'] == '2-1（16.0%） / 1-0（13.0%） / 2-0'
    assert result.loc[0, '半全场参考'] == '胜胜（36.0%） / 平胜（22.0%）'


def test_yesterday_recommendation_review_scores_only_the_primary_option(
        monkeypatch, tmp_path):
    report = pd.DataFrame([{
        '赛事编号': '周四001', '比赛时间': '2026-08-27 20:00',
        '联赛': '英超', '主队': '阿森纳', '客队': '切尔西',
        '建议状态': '精选主推', '盘口门控': '稳定',
        '胜平负首选': '胜', '胜平负首选概率': 0.66,
        '模型主胜概率': 0.66, '模型平局概率': 0.22, '模型客胜概率': 0.12,
        '首次采集胜奖金': 1.90, '首次采集平奖金': 3.40, '首次采集负奖金': 4.20,
        '官方胜奖金': 1.80, '官方平奖金': 3.50, '官方负奖金': 4.40,
        '模拟胜负': '胜 68.0%',
        '官方让球数': -1, '让球首选': '胜', '让球首选概率': 0.64,
        '模型让胜概率': 0.64, '模型让平概率': 0.20, '模型让负概率': 0.16,
        '首次采集让球数': -1,
        '首次采集让胜奖金': 2.10, '首次采集让平奖金': 3.10,
        '首次采集让负奖金': 3.20,
        '官方让胜奖金': 2.0, '官方让平奖金': 3.2, '官方让负奖金': 3.4,
        '模拟让球': '让胜 65.0%',
        '让球最大概率优势': 0.05,
        '大小球首选': '大于2.5球', '大小球首选概率': 0.61,
        '模拟竞彩总进球': '3球', '模拟竞彩总进球概率': 0.27,
        '半全场首选': '胜胜', '半全场首选概率': 0.36,
        '比分推荐状态': '推荐', '首选比分': '2-1',
        '原始最高概率比分概率': 0.16,
    }])
    report.to_csv(tmp_path / '2026-08-27-竞彩预测.csv', index=False)
    details = pd.DataFrame([{
        '赛事编号': '周四001', '完场比分': '2-1',
        '胜负': '胜 → 胜（命中）',
        '让球（首/次）': '首胜/次负 → 让负（次中）',
        '大小球': '大 → 大（命中）',
        '半全场（首/次）': '首胜胜/次平胜 → 胜胜（首中）',
        '比分（首/次1/次2/冷/进）': '2-1/1-0/2-0 → 2-1（首中）',
    }])
    monkeypatch.setattr(sporttery_window, 'REPORT_ROOT', tmp_path)
    monkeypatch.setattr(
        sporttery_window, 'load_yesterday_hit_report',
        lambda: (details, {'date': '2026-08-27'}),
    )
    frozen = sporttery_window.build_daily_recommendations(report, future_only=False)
    monkeypatch.setattr(
        sporttery_window, '_load_daily_recommendation_snapshot',
        lambda day: frozen if day == '2026-08-27' else frozen.iloc[0:0],
    )
    result, review_date = sporttery_window.build_yesterday_recommendation_review()
    assert review_date == '2026-08-27'
    statuses = dict(zip(result['推荐玩法'], result['命中状态']))
    assert statuses == {'让球胜平负': '✕ 未中'}


def test_yesterday_recommendation_review_uses_lottery_card_date(
        monkeypatch, tmp_path):
    report = pd.DataFrame([{
        '赛事编号': '周五001', '比赛时间': '2026-08-29 01:30',
        '联赛': '英超', '主队': '主队', '客队': '客队',
    }])
    report.to_csv(tmp_path / '2026-08-28-竞彩预测.csv', index=False)
    details = pd.DataFrame([{
        '赛事编号': '周五001', '完场比分': '1-0',
        '胜负': '胜 → 胜（命中）',
    }])
    selected = pd.DataFrame([{
        '赛事编号': '周五001', '联赛': '英超', '对阵': '主队 vs 客队',
        '推荐玩法': '胜负', '重点选项': '★ 胜', '模型概率': '60.0%',
    }])
    monkeypatch.setattr(sporttery_window, 'REPORT_ROOT', tmp_path)
    monkeypatch.setattr(
        sporttery_window, 'load_yesterday_hit_report',
        lambda: (details, {'date': '2026-08-28'}),
    )
    monkeypatch.setattr(
        sporttery_window, '_load_daily_recommendation_snapshot',
        lambda day: selected if day == '2026-08-28' else selected.iloc[0:0],
    )

    result, review_date = sporttery_window.build_yesterday_recommendation_review()

    assert review_date == '2026-08-28'
    assert result['赛事编号'].tolist() == ['周五001']
    assert result['命中状态'].tolist() == ['✓ 命中']


def test_export_view_includes_opening_market_information():
    predictions = pd.DataFrame([{
        '赛事编号': '周六001', '主队': '主队', '客队': '客队',
        '首次采集胜奖金': 2.10, '首次采集平奖金': 3.20, '首次采集负奖金': 3.40,
        '官方胜奖金': 2.00, '官方平奖金': 3.25, '官方负奖金': 3.55,
        '首次采集让球数': -1, '官方让球数': -1,
        '首次采集让胜奖金': 4.20, '官方让胜奖金': 4.00,
    }])

    display = sporttery_window.SportteryPredictionsDialog._display_predictions(
        predictions, include_market_details=True,
    )

    assert '胜平负指数（初盘/首次采集→当前）' in display.columns
    assert display.loc[0, '胜平负指数（初盘/首次采集→当前）'] == (
        '初 2.10 → 现 2.00｜初 3.20 → 现 3.25｜初 3.40 → 现 3.55'
    )
    assert '线 初 -1 → 现 -1' in display.loc[
        0, '让球指数（初盘/首次采集→当前）'
    ]


def test_simulation_filter_includes_low_confidence_prior_rows():
    predictions = pd.DataFrame([
        {'赛事编号': '周六001', '模拟次数': 10_000, '模拟胜负': '胜 45.0%',
         '模拟模型来源': '本地跨联赛真实比分先验（低置信）'},
        {'赛事编号': '周六002', '模拟次数': 0, '模拟胜负': '',
         '模拟模型来源': '历史攻防样本不足'},
    ])

    result = filter_predictions_by_model(predictions, SIMULATION_MODELS)

    assert result['赛事编号'].tolist() == ['周六001']


def test_sporttery_table_keeps_only_compact_decision_columns():
    predictions = pd.DataFrame([{
        '赛事编号': '周六009', '联赛': '测试联赛', '主队': '主队', '客队': '客队',
        '胜负模型类别': '市场基线', '建议状态': '精选主推', '置信等级': '高',
        '模型主胜概率': 0.70, '模型平局概率': 0.18, '模型客胜概率': 0.12,
        '胜平负首选': '胜', '胜平负首选概率': 0.70,
        '同阈值历史命中率': 0.762, '同阈值历史覆盖率': 0.137,
        '筛选回测样本数': 137, '预测依据': '官方赔率市场基线',
    }])

    display = sporttery_window.SportteryPredictionsDialog._display_predictions(predictions)

    assert display.columns.tolist() == [
        '赛事编号', '联赛', '对阵', '距参考截止',
        '推荐性质', '胜负模型', '数据状态',
        '综合方向', '盘口流向',
        '让球', '总进球', '半全场', '比分', '风险提示',
    ]
    assert display.loc[0, '对阵'] == '主队 vs 客队'
    assert display.loc[0, '综合方向'] == '胜（70.0%）'
    assert display.loc[0, '推荐性质'] == '正式主推'
    assert display.loc[0, '数据状态'] == '历史报告未标记来源'
    assert display.loc[0, '风险提示'] == '正常'


def test_daily_priority_matches_unified_daily_recommendation_cells(monkeypatch):
    predictions = pd.DataFrame([
        {
            '赛事编号': '周六001', '比赛时间': '2026-08-29 18:00',
            '建议状态': '高置信主推', '胜平负首选概率': 0.66,
            '让球首选概率': 0.64, '让球最大概率优势': 0.05,
            '大小球首选概率': 0.61, '半全场首选概率': 0.36,
            '大小球首选': '大于2.5球',
            '比分推荐状态': '推荐', '原始最高概率比分概率': 0.13,
        },
        {
            '赛事编号': '周六002', '比赛时间': '2026-08-29 20:00',
            '建议状态': '精选主推', '胜平负首选概率': 0.72,
            '让球首选概率': 0.68, '让球最大概率优势': 0.06,
            '大小球首选概率': 0.63, '半全场首选概率': 0.38,
            '大小球首选': '大于2.5球',
            '比分推荐状态': '推荐', '原始最高概率比分概率': 0.14,
        },
    ])

    monkeypatch.setattr(
        sporttery_window, 'build_daily_recommendations',
        lambda frame, future_only=False: pd.DataFrame([{
            '赛事编号': '周六002', '推荐玩法': '比分',
        }]),
    )
    priorities = sporttery_window._daily_priority_aspects(predictions)

    assert priorities.iloc[0] == []
    assert priorities.iloc[1] == ['比分']


def test_daily_priority_rejects_unstable_market_signals():
    predictions = pd.DataFrame([{
        '比赛时间': '2026-08-29 20:00', '建议状态': '精选主推',
        '胜平负首选概率': 0.80, '让球首选概率': 0.75,
        '让球最大概率优势': 0.10, '大小球首选概率': 0.70,
        '半全场首选概率': 0.45, '比分推荐状态': '推荐',
        '大小球首选': '大于2.5球',
        '原始最高概率比分概率': 0.16, '盘口门控': '盘口震荡不稳定',
    }])

    assert sporttery_window._daily_priority_aspects(predictions).iloc[0] == []


def test_priority_summary_only_lists_markets_that_pass_the_gate(monkeypatch):
    predictions = pd.DataFrame({'赛事编号': ['001', '002']})
    monkeypatch.setattr(
        sporttery_window,
        '_daily_priority_aspects',
        lambda frame: pd.Series([['胜负', '比分'], ['总进球']]),
    )
    assert sporttery_window._priority_summary(predictions) == (
        '今日重点 胜负1·总进球1·比分1'
    )


def test_daily_recommendation_displays_lottery_card_day_after_midnight(monkeypatch):
    predictions = pd.DataFrame([{
        '赛事编号': '周日018', '比赛时间': '2026-08-31 01:00',
        '联赛': '荷甲', '主队': '甲', '客队': '乙',
        '盘口门控': '稳定', '胜平负首选': '胜', '胜平负首选概率': .68,
        '模型主胜概率': .68, '模型平局概率': .20, '模型客胜概率': .12,
        '首次采集胜奖金': 1.90, '首次采集平奖金': 3.40,
        '首次采集负奖金': 4.20, '官方胜奖金': 1.80,
        '官方平奖金': 3.50, '官方负奖金': 4.40,
        '模拟胜负': '胜 69.0%',
    }])
    monkeypatch.setattr(
        sporttery_window, '_daily_priority_aspects',
        lambda frame: pd.Series([['大小球']]),
    )

    result = sporttery_window.build_daily_recommendations(
        predictions, future_only=False,
    )

    assert result.loc[0, '比赛日期'] == '2026-08-30'
    assert result.loc[0, '赛事编号'] == '周日018'
    assert result.loc[0, '推荐玩法'] == '胜平负'
    assert result.loc[0, '重点选项'] == '★ 胜'


def test_daily_recommendation_uses_dual_pick_only_when_draw_protection_triggered():
    predictions = pd.DataFrame([{
        '赛事编号': '周三001', '比赛时间': '2099-09-02 19:00',
        '联赛': '测试联赛', '主队': '甲', '客队': '乙',
        '盘口门控': '稳定', '胜平负首选': '胜', '胜平负首选概率': .36,
        '平局双选保护': '胜平',
        '模型主胜概率': .36, '模型平局概率': .30, '模型客胜概率': .34,
        '首次采集胜奖金': 2.70, '首次采集平奖金': 3.20,
        '首次采集负奖金': 2.90, '官方胜奖金': 2.65,
        '官方平奖金': 3.20, '官方负奖金': 3.00,
        '模拟胜负': '平 35.0%',
    }])
    result = sporttery_window.build_daily_recommendations(
        predictions, future_only=False,
    )
    assert result.loc[0, '推荐玩法'] == '胜平负双选'
    assert result.loc[0, '重点选项'] == '· 胜平'
    assert result.loc[0, '推荐等级'] == '平局双选保护'
    assert result.loc[0, '建议仓位'] == '仅作平局保护，不计单选仓位'


def test_daily_recommendations_fill_card_with_explicit_observations():
    predictions = pd.DataFrame([{
        '赛事编号': f'周日{index:03d}', '比赛时间': f'2099-08-30 {10 + index:02d}:00',
        '联赛': '测试联赛', '主队': f'主{index}', '客队': f'客{index}',
        '盘口门控': '稳定', '胜平负首选': '胜',
        '胜平负首选概率': 0.55 + index / 100,
        '模型主胜概率': 0.55 + index / 100,
        '模型平局概率': 0.27 - index / 200,
        '模型客胜概率': 0.18 - index / 200,
        '首次采集胜奖金': 1.90, '首次采集平奖金': 3.40,
        '首次采集负奖金': 4.20, '官方胜奖金': 1.80,
        '官方平奖金': 3.50, '官方负奖金': 4.40,
        '模拟胜负': '胜 62.0% / 平 23.0%',
    } for index in range(1, 11)])

    result = sporttery_window.build_daily_recommendations(
        predictions, future_only=False,
    )

    assert len(result) == 5
    assert result['相对安全等级'].str.match(r'[A-E]｜').all()
    assert result.loc[
        result['推荐性质'].eq('观察/不投注'), '行动结论'
    ].eq('不建议投注').all()
    assert result['赛事编号'].nunique() == len(result)
    assert result['推荐玩法'].eq('胜平负').all()
    assert set(result['推荐等级']) <= {'核心重点', '可买优选', '综合观察'}
    assert result['赛事编号'].tolist() == [
        '周日010', '周日009', '周日008', '周日007', '周日006',
    ]
    assert result['蒙特卡洛是否同向'].str.startswith('同向').all()
    assert result['盘口验证'].str.contains('支持|走强').all()
    assert result.loc[
        result['推荐等级'].eq('综合观察'), '建议仓位'
    ].eq('不投注').all()


def test_daily_recommendations_keep_best_five_per_day():
    predictions = pd.DataFrame([{
        '赛事编号': f'{weekday}{index:03d}',
        '比赛时间': f'2099-08-{day:02d} {10 + index:02d}:00',
        '联赛': '测试联赛', '主队': f'主{day}-{index}',
        '客队': f'客{day}-{index}', '盘口门控': '稳定',
        '胜平负首选': '胜', '胜平负首选概率': 0.54 + index / 100,
        '模型主胜概率': 0.54 + index / 100,
        '模型平局概率': 0.28 - index / 200,
        '模型客胜概率': 0.18 - index / 200,
        '首次采集胜奖金': 1.90, '首次采集平奖金': 3.40,
        '首次采集负奖金': 4.20, '官方胜奖金': 1.80,
        '官方平奖金': 3.50, '官方负奖金': 4.40,
        '模拟胜负': '胜 62.0% / 平 23.0%',
    } for day, weekday in ((29, '周六'), (30, '周日'), (31, '周一'))
      for index in range(1, 8)])

    result = sporttery_window.build_daily_recommendations(
        predictions, future_only=False,
    )

    assert len(result) == 15
    assert result[['比赛日期', '赛事编号']].drop_duplicates().shape[0] == 15
    assert result['比赛日期'].value_counts().to_dict() == {
        '2099-08-29': 5,
        '2099-08-30': 5,
        '2099-08-31': 5,
    }
    assert result.iloc[:5]['比赛日期'].eq('2099-08-29').all()


def test_daily_recommendations_reject_retained_stopped_snapshot():
    predictions = pd.DataFrame([{
        '赛事编号': '周日001', '比赛时间': '2099-08-30 18:00',
        '联赛': '测试联赛', '主队': '主队', '客队': '客队',
        '官方销售状态': '已退出当前在售列表', '同步时段': '停止推荐',
        '盘口门控': '稳定', '胜平负首选': '胜', '胜平负首选概率': .70,
        '模型主胜概率': .70, '模型平局概率': .18, '模型客胜概率': .12,
        '首次采集胜奖金': 1.90, '首次采集平奖金': 3.40,
        '首次采集负奖金': 4.20, '官方胜奖金': 1.80,
        '官方平奖金': 3.50, '官方负奖金': 4.40,
        '模拟胜负': '胜 72.0%', '胜负模型类别': '英超专用模型',
    }])

    assert sporttery_window.build_daily_recommendations(
        predictions, future_only=False,
    ).empty


def test_daily_recommendations_reject_high_lineup_warning():
    predictions = pd.DataFrame([{
        '赛事编号': '周日001', '比赛时间': '2099-08-30 18:00',
        '联赛': '测试联赛', '主队': '主队', '客队': '客队',
        '盘口门控': '稳定', '胜平负首选': '胜', '胜平负首选概率': .68,
        '模型主胜概率': .68, '模型平局概率': .20, '模型客胜概率': .12,
        '首次采集胜奖金': 1.90, '首次采集平奖金': 3.40,
        '首次采集负奖金': 4.20, '官方胜奖金': 1.80,
        '官方平奖金': 3.50, '官方负奖金': 4.40,
        '模拟胜负': '胜 69.0%', '首发状态': '已确认',
        '阵容预警级别': '高', '阵容方向冲突': False,
    }])
    assert sporttery_window.build_daily_recommendations(
        predictions, future_only=False,
    ).empty


def test_daily_recommendations_reject_monte_or_market_conflict():
    base = {
        '比赛时间': '2099-08-30 18:00', '联赛': '测试联赛',
        '主队': '主队', '客队': '客队', '盘口门控': '稳定',
        '胜平负首选': '胜', '胜平负首选概率': .68,
        '模型主胜概率': .68, '模型平局概率': .20, '模型客胜概率': .12,
        '首次采集胜奖金': 1.90, '首次采集平奖金': 3.40,
        '首次采集负奖金': 4.20,
    }
    predictions = pd.DataFrame([
        {
            **base, '赛事编号': '周日001', '模拟胜负': '负 55.0%',
            '官方胜奖金': 1.80, '官方平奖金': 3.50, '官方负奖金': 4.40,
        },
        {
            **base, '赛事编号': '周日002', '模拟胜负': '胜 65.0%',
            '官方胜奖金': 3.20, '官方平奖金': 3.00, '官方负奖金': 2.10,
        },
    ])

    result = sporttery_window.build_daily_recommendations(
        predictions, future_only=False,
    )

    assert result.empty


def test_daily_recommendations_do_not_promote_score_or_half_full():
    predictions = pd.DataFrame([{
        '赛事编号': '周日001', '比赛时间': '2099-08-30 18:00',
        '联赛': '测试联赛', '主队': '主队', '客队': '客队',
        '比分推荐状态': '推荐', '首选比分': '2-1', '首选比分概率': .30,
        '模拟Top3比分': '2-1 30.0% / 1-0 20.0%',
        '半全场首选': '胜胜', '半全场首选概率': .70,
        '模拟半全场': '胜胜 72.0%',
    }])

    result = sporttery_window.build_daily_recommendations(
        predictions, future_only=False,
    )

    assert result.empty


def test_daily_recommendations_keep_low_value_handicap_as_unstarred_observation():
    predictions = pd.DataFrame([{
        '赛事编号': '周日002', '比赛时间': '2099-08-30 19:00',
        '联赛': '测试联赛', '主队': '主队', '客队': '客队',
        '盘口门控': '稳定', '官方让球数': -1,
        '让球首选': '负', '让球首选概率': .60,
        '模型让胜概率': .18, '模型让平概率': .22, '模型让负概率': .60,
        '首次采集让球数': -1,
        '首次采集让胜奖金': 4.20, '首次采集让平奖金': 3.40,
        '首次采集让负奖金': 1.50,
        '官方让胜奖金': 4.20, '官方让平奖金': 3.40,
        '官方让负奖金': 1.50, '模拟让球': '让负 61.0%',
    }])

    result = sporttery_window.build_daily_recommendations(
        predictions, future_only=False,
    )

    assert len(result) == 1
    assert result.loc[0, '推荐等级'] == '盘口观察'
    assert result.loc[0, '重点选项'] == '· -1球 负'
    assert result.loc[0, '建议仓位'] == '不投注'


def test_daily_recommendations_keep_best_score_without_parlay_fields():
    predictions = pd.DataFrame([{
        '赛事编号': '周日003', '比赛时间': '2099-08-30 20:00',
        '联赛': '测试联赛', '主队': '主队', '客队': '客队',
        '盘口门控': '稳定', '胜平负首选': '胜', '胜平负首选概率': .62,
        '模型主胜概率': .62, '模型平局概率': .15, '模型客胜概率': .23,
        '首次采集胜奖金': 1.80, '首次采集平奖金': 3.40,
        '首次采集负奖金': 4.70, '官方胜奖金': 1.80,
        '官方平奖金': 3.40, '官方负奖金': 4.70,
        '模拟胜负': '胜 63.0%', '首选比分': '2-1', '首选比分概率': .16,
        '次选比分': '1-0', '次选比分概率': .13,
    }])

    result = sporttery_window.build_daily_recommendations(
        predictions, future_only=False,
    )

    assert len(result) == 1
    assert result.loc[0, '最佳比分'] == '◎ 2-1（16.0%）'
    assert not {'每日2串1', '2串1组合概率', '2串1组合SP', '阵容验证'} & set(result.columns)
